import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, Tuple

import pandas as pd


DEFAULT_DATABASES = ["ACM", "IEEE", "Scopus", "ERIC", "ScienceDirect"]

COLUMN_ALIASES = {
    "Title": ["title", "document title", "article title", "publication title"],
    "Abstract": ["abstract", "abstract note", "description"],
    "Authors": ["authors", "author", "author names"],
    "Year": ["year", "publication year", "date"],
    "DOI": ["doi", "digital object identifier"],
    "URL": ["url", "link", "document url", "article url"],
    "PDF_URL": ["pdf_url", "pdf url", "full text url", "fulltext url", "download url"],
    "Source_Title": ["source title", "journal", "publication name", "booktitle", "proceedings title"],
}

TITLE_INCLUDE_KEYWORDS = [
    "computational thinking",
    "programming",
    "coding",
    "computer science education",
    "scratch",
    "block based",
    "block-based",
    "python",
    "robotics",
    "physical computing",
    "algorithmic thinking",
]

TITLE_PARTIAL_KEYWORDS = [
    "problem solving",
    "stem",
    "digital literacy",
    "computing",
    "algorithm",
    "informatics",
    "educational technology",
    "learning analytics",
]

TITLE_EXCLUDE_KEYWORDS = [
    "medical image",
    "clinical",
    "hospital",
    "diagnosis",
    "protein",
    "genomics",
    "bioinformatics",
    "neuroscience",
    "radiology",
    "disease",
]

NON_STUDY_PATTERNS = {
    "Proceedings volume": [
        r"^.*proceedings of the .*conference.*$",
        r"^.*proceedings of .*conference.*$",
        r"^.*conference proceedings.*$",
        r"^.*\bproceedings\b.*\b\d{4}\b.*$",
        r"^.*\bUKICER\b.*\bproceedings\b.*$",
        r"^.*\bSIGCSE\b.*\bproceedings\b.*$",
        r"^.*\bCHI\b.*\bproceedings\b.*$",
        r"^.*\bIDC\b.*\bproceedings\b.*$",
        r"^.*\bsigcsets?\b.*\bproceedings\b.*$",
        r"^.*\bproceedings\b.*\bacm technical symposium\b.*$",
        r"^.*\bproceedings\b.*\bcomputer science education\b.*v\.?\s*\d+.*$",
        r"^.*\bproceedings\b.*\b\d+(st|nd|rd|th)\s+acm technical symposium\b.*$",
        r"^.*sigcsets?\s*\d{4}\s*:\s*proceedings.*$",
        r".*\bproceedings\b.*\bsymposium\b.*",
        r".*\bproceedings\b.*\btechnical symposium\b.*",
        r".*\bproceedings of the .*symposium\b.*",
    ],
    "Proceedings front matter": [
        r"\bfront matter\b",
        r"\bwelcome message\b",
        r"\bmessage from the chair\b",
        r"\bcommittee\b",
        r"\borganizing committee\b",
        r"\bprogram committee\b",
        r"\btable of contents\b",
        r"\bauthor index\b",
        r"\bindex\b",
        r"\bcontents\b",
        r"\bsession overview\b",
        r"\bconference overview\b",
        r"\bconference program\b",
        r"\bproceedings preface\b",
    ],
}

PUBLICATION_TYPE_PATTERNS = {
    "Poster candidate": [
        r"\bposter\b",
        r"\bposter abstract\b",
        r"\bposter presentation\b",
    ],
    "Demo candidate": [
        r"\bdemo\b",
        r"\bdemonstration\b",
    ],
    "WIP candidate": [
        r"\bwork in progress\b",
        r"\bwip\b",
        r"\blate breaking work\b",
        r"\blbw\b",
    ],
    "Extended abstract candidate": [
        r"\bextended abstract\b",
        r"\bextended abstracts\b",
        r"\bcompanion proceedings\b",
        r"\bconference companion\b",
    ],
}


# ============================================================
# Text cleaning
# ============================================================

def clean_excel_value(value: Any) -> Any:
    """Remove characters that can break Excel XML files."""
    if pd.isna(value):
        return ""

    if not isinstance(value, str):
        return value

    value = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)
    value = value.replace("\uFFFE", "").replace("\uFFFF", "")

    if len(value) > 32000:
        value = value[:32000]

    return value


def sanitize_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Prepare a dataframe for safe Excel export."""
    if df is None:
        return pd.DataFrame()

    out = df.copy()

    for col in out.columns:
        out[col] = out[col].map(clean_excel_value)

    return out


def clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""

    text = str(value).lower().strip()
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("utf-8")
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def clean_title(title: Any) -> str:
    return clean_text(title)


def clean_abstract(abstract: Any) -> str:
    return clean_text(abstract)


def clean_authors(authors: Any) -> str:
    return clean_text(authors)


def clean_doi(doi: Any) -> str:
    if pd.isna(doi):
        return ""

    text = str(doi).lower().strip()
    text = text.replace("https://doi.org/", "")
    text = text.replace("http://doi.org/", "")
    text = text.replace("doi:", "")
    text = text.strip().strip(".")

    return text


# ============================================================
# Input standardization
# ============================================================

def infer_query_type(sheet_name: Any) -> str:
    name = str(sheet_name).lower()

    has_evaluation = any(term in name for term in ["evaluation", "evaluating", "assessment"])
    has_intervention = any(term in name for term in ["intervention", "teaching", "learning"])

    if has_evaluation and has_intervention:
        return "evaluation; intervention"

    if has_evaluation:
        return "evaluation"

    if has_intervention:
        return "intervention"

    return "unknown"


def find_column(df: pd.DataFrame, standard_name: str) -> str | None:
    aliases = COLUMN_ALIASES.get(standard_name, [])

    normalized_columns = {
        str(col).strip().lower(): col
        for col in df.columns
    }

    for alias in aliases:
        if alias in normalized_columns:
            return normalized_columns[alias]

    return None


def standardize_sheet(df: pd.DataFrame, source_database: str, sheet_name: str) -> pd.DataFrame:
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    standardized = pd.DataFrame()

    for standard_col in COLUMN_ALIASES:
        source_col = find_column(df, standard_col)
        standardized[standard_col] = df[source_col] if source_col is not None else ""

    standardized["Source_Database"] = source_database
    standardized["Sheet_Name"] = sheet_name
    standardized["Query_Type"] = infer_query_type(sheet_name)
    standardized["Original_Row_Number"] = range(1, len(standardized) + 1)

    standardized["Title_Clean"] = standardized["Title"].apply(clean_title)
    standardized["Abstract_Clean"] = standardized["Abstract"].apply(clean_abstract)
    standardized["Authors_Clean"] = standardized["Authors"].apply(clean_authors)
    standardized["DOI_Clean"] = standardized["DOI"].apply(clean_doi)

    return standardized


def read_database_workbook(uploaded_file, source_database: str) -> pd.DataFrame:
    sheets = pd.read_excel(uploaded_file, sheet_name=None, dtype=str)

    standardized_sheets = []

    for sheet_name, df in sheets.items():
        if df.empty:
            continue

        standardized = standardize_sheet(
            df=df,
            source_database=source_database,
            sheet_name=sheet_name,
        )
        standardized_sheets.append(standardized)

    if not standardized_sheets:
        return pd.DataFrame()

    return pd.concat(standardized_sheets, ignore_index=True)


def assign_record_ids(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    if "Record_ID" not in out.columns:
        out.insert(0, "Record_ID", [f"R{i:05d}" for i in range(1, len(out) + 1)])

    return out


# ============================================================
# Deduplication
# ============================================================

def calculate_completeness_score(row: pd.Series) -> int:
    score = 0

    if str(row.get("DOI_Clean", "")).strip():
        score += 5

    if str(row.get("Abstract", "")).strip():
        score += 3

    if len(str(row.get("Abstract", "")).strip()) >= 300:
        score += 2

    if str(row.get("Authors", "")).strip():
        score += 2

    if str(row.get("Year", "")).strip():
        score += 1

    if str(row.get("URL", "")).strip():
        score += 1

    if str(row.get("PDF_URL", "")).strip():
        score += 2

    if str(row.get("Source_Title", "")).strip():
        score += 1

    return score


def build_exact_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    rows = []

    for key_col, duplicate_type in [
        ("DOI_Clean", "DOI exact match"),
        ("Title_Clean", "Title exact match"),
        ("Abstract_Clean", "Abstract exact match"),
    ]:
        if key_col not in df.columns:
            continue

        valid = df[df[key_col].astype(str).str.strip() != ""]
        groups = valid.groupby(key_col)["Record_ID"].apply(list).reset_index()

        for _, row in groups.iterrows():
            record_ids = row["Record_ID"]

            if len(record_ids) > 1:
                rows.append(
                    {
                        "Duplicate_Type": duplicate_type,
                        "Match_Value": row[key_col],
                        "Record_IDs": "; ".join(record_ids),
                        "Count": len(record_ids),
                    }
                )

    return pd.DataFrame(rows)


def mark_exact_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    out["Duplicate_Status"] = "Unique"
    out["Duplicate_Group_ID"] = ""
    out["Duplicate_Reason"] = ""
    out["Completeness_Score"] = out.apply(calculate_completeness_score, axis=1)

    group_counter = 1

    for key_col, reason in [
        ("DOI_Clean", "DOI exact match"),
        ("Title_Clean", "Title exact match"),
        ("Abstract_Clean", "Abstract exact match"),
    ]:
        if key_col not in out.columns:
            continue

        valid = out[
            (out[key_col].astype(str).str.strip() != "")
            & (out["Duplicate_Status"] != "Duplicate")
        ]

        groups = valid.groupby(key_col)["Record_ID"].apply(list)

        for _, record_ids in groups.items():
            if len(record_ids) <= 1:
                continue

            group_id = f"D{group_counter:04d}"
            group_counter += 1

            group_df = out[out["Record_ID"].isin(record_ids)].copy()

            keep_id = (
                group_df
                .sort_values(by=["Completeness_Score", "Record_ID"], ascending=[False, True])
                ["Record_ID"]
                .iloc[0]
            )

            remove_ids = [rid for rid in record_ids if rid != keep_id]

            out.loc[out["Record_ID"].isin(record_ids), "Duplicate_Group_ID"] = group_id
            out.loc[out["Record_ID"].isin(record_ids), "Duplicate_Reason"] = reason
            out.loc[out["Record_ID"] == keep_id, "Duplicate_Status"] = "Retained"
            out.loc[out["Record_ID"].isin(remove_ids), "Duplicate_Status"] = "Duplicate"

    return out


def build_fuzzy_duplicates() -> pd.DataFrame:
    """Create an empty audit table because expensive fuzzy checks are deferred to abstract screening."""
    return pd.DataFrame(
        columns=[
            "Record_ID_A",
            "Title_A",
            "Abstract_A",
            "Authors_A",
            "Year_A",
            "DOI_A",
            "Record_ID_B",
            "Title_B",
            "Abstract_B",
            "Authors_B",
            "Year_B",
            "DOI_B",
            "Title_Similarity",
            "Abstract_Similarity",
            "Author_Similarity",
            "Combined_Similarity",
            "Suggested_Action",
            "Manual_Duplicate_Decision",
            "Manual_Notes",
        ]
    )


# ============================================================
# Screenability and title screening
# ============================================================

def get_screenability_flag(row: pd.Series) -> Tuple[str, str]:
    title = str(row.get("Title", "")).lower()
    source = str(row.get("Source_Title", "")).lower()

    for label, patterns in NON_STUDY_PATTERNS.items():
        text_to_check = title if label == "Proceedings volume" else f"{title} {source}"

        for pattern in patterns:
            if re.search(pattern, text_to_check, flags=re.IGNORECASE):
                return "Exclude before title screening", f"{label}: matched pattern {pattern}"

    return "Screenable", ""

def get_publication_type_flag(row: pd.Series) -> Tuple[str, str]:
    title = str(row.get("Title", ""))
    abstract = str(row.get("Abstract", ""))
    source = str(row.get("Source_Title", ""))
    combined = f"{title} {abstract} {source}".lower()

    for label, patterns in PUBLICATION_TYPE_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, combined, flags=re.IGNORECASE):
                return label, f"Matched pattern: {pattern}"

    return "Regular paper candidate", ""


def get_abstract_quality_flag(row: pd.Series) -> Tuple[str, str]:
    abstract = str(row.get("Abstract", "")).strip()
    abstract_lower = abstract.lower()
    word_count = len(abstract.split())

    if abstract_lower in ["", "nan", "none", "null", "no abstract", "no abstract available", "not available", "n/a", "na"]:
        return "Missing abstract", "No usable abstract text available"

    if word_count < 40:
        return "Short abstract", f"Only {word_count} words"

    truncated_endings = [
        r"\band\s*$",
        r"\bor\s*$",
        r"\bwith\s*$",
        r"\bfrom\s*$",
        r"\bin\s*$",
        r"\bthe\s*$",
        r"\bof\s*$",
        r"\bto\s*$",
    ]

    if any(re.search(pattern, abstract_lower) for pattern in truncated_endings):
        return "Possibly truncated abstract", "Abstract ends with incomplete phrase"

    return "Usable abstract", ""

def apply_screenability_labels(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    screenability = out.apply(get_screenability_flag, axis=1, result_type="expand")
    out["Screenability_Flag"] = screenability[0]
    out["Screenability_Reason"] = screenability[1]

    publication_type = out.apply(get_publication_type_flag, axis=1, result_type="expand")
    out["Publication_Type_Flag"] = publication_type[0]
    out["Publication_Type_Reason"] = publication_type[1]

    abstract_quality = out.apply(get_abstract_quality_flag, axis=1, result_type="expand")
    out["Abstract_Quality_Flag"] = abstract_quality[0]
    out["Abstract_Quality_Reason"] = abstract_quality[1]

    return out


def automatic_title_screening(row: pd.Series) -> Tuple[str, str]:
    title = str(row.get("Title", "")).strip()
    title_clean = clean_title(title)

    if not title_clean:
        return "Undetermined", "Missing or unreadable title"

    if any(term in title_clean for term in TITLE_EXCLUDE_KEYWORDS):
        return "No", "Title contains clear exclusion keyword"

    include_hits = [term for term in TITLE_INCLUDE_KEYWORDS if term in title_clean]
    partial_hits = [term for term in TITLE_PARTIAL_KEYWORDS if term in title_clean]

    if len(include_hits) >= 2:
        return "Yes", "Title contains multiple strong inclusion keywords"

    if "computational thinking" in include_hits and partial_hits:
        return "Yes", "Title contains computational thinking and education-related context"

    if include_hits:
        return "Partially", "Title contains one strong inclusion keyword"

    if partial_hits:
        return "Partially", "Title contains potentially relevant keyword"

    return "No", "Title does not indicate the main scope of the review"


def apply_title_screening(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    decisions = out.apply(automatic_title_screening, axis=1, result_type="expand")
    out["Title_Screening_Decision"] = decisions[0]
    out["Title_Screening_Reason"] = decisions[1]
    out["Title_Reviewer_Notes"] = ""

    return out


def build_clean_master_with_labels(deduplicated_df: pd.DataFrame) -> pd.DataFrame:
    clean = deduplicated_df[deduplicated_df["Duplicate_Status"] != "Duplicate"].copy()
    clean = apply_screenability_labels(clean)

    clean["Manual_Review_Label"] = ""
    clean.loc[clean["Screenability_Flag"] != "Screenable", "Manual_Review_Label"] = clean.loc[
        clean["Screenability_Flag"] != "Screenable",
        "Screenability_Flag",
    ]

    return clean


def build_non_screenable_records(clean_master: pd.DataFrame) -> pd.DataFrame:
    if clean_master.empty or "Screenability_Flag" not in clean_master.columns:
        return pd.DataFrame()

    return clean_master[clean_master["Screenability_Flag"] != "Screenable"].copy()


def build_manual_review_needed(clean_master: pd.DataFrame) -> pd.DataFrame:
    if clean_master.empty or "Manual_Review_Label" not in clean_master.columns:
        return pd.DataFrame()

    return clean_master[clean_master["Manual_Review_Label"].astype(str).str.strip() != ""].copy()


def build_title_screening_results(clean_master: pd.DataFrame) -> pd.DataFrame:
    if clean_master.empty:
        return pd.DataFrame()

    screenable = clean_master[
        clean_master["Screenability_Flag"] == "Screenable"
    ].copy()

    return apply_title_screening(screenable)


def build_title_screening_summary(title_screening_results: pd.DataFrame) -> pd.DataFrame:
    if title_screening_results.empty:
        return pd.DataFrame(columns=["Title_Screening_Decision", "Count"])

    summary = (
        title_screening_results["Title_Screening_Decision"]
        .value_counts(dropna=False)
        .reset_index()
    )

    summary.columns = ["Title_Screening_Decision", "Count"]

    total = pd.DataFrame(
        [{"Title_Screening_Decision": "TOTAL", "Count": int(summary["Count"].sum())}]
    )

    return pd.concat([summary, total], ignore_index=True)


def build_abstract_screening_input(title_screening_results: pd.DataFrame) -> pd.DataFrame:
    if title_screening_results.empty:
        return pd.DataFrame()

    keep = title_screening_results[
        title_screening_results["Title_Screening_Decision"].isin(["Yes", "Partially"])
    ].copy()

    preferred_columns = [
        "Record_ID",
        "Title",
        "Abstract",
        "Authors",
        "Year",
        "DOI",
        "URL",
        "PDF_URL",
        "Source_Title",
        "Source_Database",
        "Sheet_Name",
        "Query_Type",
        "Duplicate_Status",
        "Duplicate_Group_ID",
        "Duplicate_Reason",
        "Screenability_Flag",
        "Screenability_Reason",
        "Manual_Review_Label",
        "Title_Screening_Decision",
        "Title_Screening_Reason",
        "Title_Reviewer_Notes",
        "Title_Clean",
        "Abstract_Clean",
        "DOI_Clean",
        "Publication_Type_Flag",
        "Publication_Type_Reason",
        "Abstract_Quality_Flag",
        "Abstract_Quality_Reason",
    ]

    existing = [col for col in preferred_columns if col in keep.columns]
    remaining = [col for col in keep.columns if col not in existing]

    return keep[existing + remaining]


# ============================================================
# Counts
# ============================================================

def build_database_statistics(standardized_all: pd.DataFrame) -> pd.DataFrame:
    if standardized_all.empty:
        return pd.DataFrame(columns=["Source_Database", "Query_Type", "Number_of_records"])

    stats = (
        standardized_all
        .groupby(["Source_Database", "Query_Type"])
        .size()
        .reset_index(name="Number_of_records")
    )

    total = pd.DataFrame(
        [
            {
                "Source_Database": "TOTAL",
                "Query_Type": "All",
                "Number_of_records": int(stats["Number_of_records"].sum()),
            }
        ]
    )

    return pd.concat([stats, total], ignore_index=True)


def build_prisma_counts(
    standardized_all: pd.DataFrame,
    merged_all: pd.DataFrame,
    deduplicated_marked: pd.DataFrame,
    exact_duplicates: pd.DataFrame,
    fuzzy_duplicates: pd.DataFrame,
    non_screenable_records: pd.DataFrame,
    title_screening_results: pd.DataFrame,
    abstract_screening_input: pd.DataFrame,
) -> pd.DataFrame:
    exact_removed = int((deduplicated_marked["Duplicate_Status"] == "Duplicate").sum())
    retained_after_exact = int((deduplicated_marked["Duplicate_Status"] != "Duplicate").sum())

    title_yes = 0
    title_partial = 0
    title_no = 0
    title_undetermined = 0

    poster_demo_candidates = 0
    missing_abstracts = 0
    short_abstracts = 0
    truncated_abstracts = 0

    if "Publication_Type_Flag" in title_screening_results.columns:
        poster_demo_candidates = int(
            title_screening_results["Publication_Type_Flag"]
            .astype(str)
            .str.contains("Poster|Demo|WIP|Extended abstract", case=False, na=False)
            .sum()
        )

    if "Abstract_Quality_Flag" in title_screening_results.columns:
        quality_counts = title_screening_results["Abstract_Quality_Flag"].value_counts().to_dict()
        missing_abstracts = int(quality_counts.get("Missing abstract", 0))
        short_abstracts = int(quality_counts.get("Short abstract", 0))
        truncated_abstracts = int(quality_counts.get("Possibly truncated abstract", 0))

    if not title_screening_results.empty:
        counts = title_screening_results["Title_Screening_Decision"].value_counts().to_dict()
        title_yes = int(counts.get("Yes", 0))
        title_partial = int(counts.get("Partially", 0))
        title_no = int(counts.get("No", 0))
        title_undetermined = int(counts.get("Undetermined", 0))

    rows = [
        {"Stage": "Records imported", "Count": len(standardized_all)},
        {"Stage": "Records after merge", "Count": len(merged_all)},
        {"Stage": "Exact duplicate groups identified", "Count": len(exact_duplicates)},
        {"Stage": "Exact duplicates removed", "Count": exact_removed},
        {"Stage": "Records retained after exact deduplication", "Count": retained_after_exact},
        {"Stage": "Possible fuzzy duplicates for manual review", "Count": len(fuzzy_duplicates)},
        {"Stage": "Records excluded before title screening as non-screenable", "Count": len(non_screenable_records)},
        {"Stage": "Records screened by title", "Count": len(title_screening_results)},
        {"Stage": "Title screening Yes", "Count": title_yes},
        {"Stage": "Title screening Partially", "Count": title_partial},
        {"Stage": "Title screening No", "Count": title_no},
        {"Stage": "Title screening Undetermined", "Count": title_undetermined},
        {"Stage": "Records sent to abstract screening", "Count": len(abstract_screening_input)},
        {"Stage": "Poster/demo/WIP/extended abstract candidates sent to title screening", "Count": poster_demo_candidates},
        {"Stage": "Missing abstracts sent to title screening", "Count": missing_abstracts},
        {"Stage": "Short abstracts sent to title screening", "Count": short_abstracts},
        {"Stage": "Possibly truncated abstracts sent to title screening", "Count": truncated_abstracts},
    ]

    return pd.DataFrame(rows)


# ============================================================
# Main pipeline
# ============================================================

def run_stage1_pipeline(uploaded_files_by_database: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
    standardized_frames = []

    for database, uploaded_file in uploaded_files_by_database.items():
        if uploaded_file is None:
            continue

        db_data = read_database_workbook(uploaded_file, database)

        if not db_data.empty:
            standardized_frames.append(db_data)

    if not standardized_frames:
        raise ValueError("No valid records were loaded.")

    standardized_all = pd.concat(standardized_frames, ignore_index=True)
    merged_all = assign_record_ids(standardized_all)

    exact_duplicates = build_exact_duplicates(merged_all)
    deduplicated_marked = mark_exact_duplicates(merged_all)

    fuzzy_duplicates = build_fuzzy_duplicates()

    clean_master = build_clean_master_with_labels(deduplicated_marked)
    non_screenable_records = build_non_screenable_records(clean_master)
    manual_review_needed = build_manual_review_needed(clean_master)

    title_screening_results = build_title_screening_results(clean_master)
    title_screening_summary = build_title_screening_summary(title_screening_results)
    abstract_screening_input = build_abstract_screening_input(title_screening_results)

    database_statistics = build_database_statistics(standardized_all)

    prisma_counts = build_prisma_counts(
        standardized_all=standardized_all,
        merged_all=merged_all,
        deduplicated_marked=deduplicated_marked,
        exact_duplicates=exact_duplicates,
        fuzzy_duplicates=fuzzy_duplicates,
        non_screenable_records=non_screenable_records,
        title_screening_results=title_screening_results,
        abstract_screening_input=abstract_screening_input,
    )

    return {
        "standardized_all_records": standardized_all,
        "merged_all_records": merged_all,
        "exact_duplicates": exact_duplicates,
        "deduplicated_marked": deduplicated_marked,
        "fuzzy_duplicates": fuzzy_duplicates,
        "non_screenable_records": non_screenable_records,
        "manual_review_needed": manual_review_needed,
        "clean_master_with_labels": clean_master,
        "title_screening_results": title_screening_results,
        "title_screening_summary": title_screening_summary,
        "abstract_screening_input": abstract_screening_input,
        "database_statistics": database_statistics,
        "prisma_counts": prisma_counts,
    }


# ============================================================
# Excel writers
# ============================================================

def style_excel_workbook(path: Path) -> None:
    """Apply safe formatting to a saved workbook."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    path = Path(path)

    if not path.exists():
        return

    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 10

            for cell in list(column_cells)[:100]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 60))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    wb.save(path)


def write_workbook_to_disk(path: Path, sheets: Dict[str, pd.DataFrame], styled: bool = True) -> Path:
    """Write a workbook directly to disk and validate that it can be reopened."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        path.unlink()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_sheet_name = str(sheet_name)[:31]
            safe_df = sanitize_dataframe_for_excel(df)
            safe_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

    if styled:
        style_excel_workbook(path)

    # Validate workbook integrity immediately.
    from openpyxl import load_workbook
    wb = load_workbook(path)
    wb.close()

    return path


def write_stage1_output_workbooks(results: Dict[str, pd.DataFrame], output_dir: Path) -> Dict[str, Path]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    processing_path = output_dir / "review_pipeline_processing.xlsx"
    title_path = output_dir / "title_screening_results.xlsx"
    abstract_input_path = output_dir / "abstract_screening_input.xlsx"

    write_workbook_to_disk(
        processing_path,
        {
            "01_standardized": results["standardized_all_records"],
            "02_merged": results["merged_all_records"],
            "03_exact_duplicates": results["exact_duplicates"],
            "04_fuzzy_duplicates": results["fuzzy_duplicates"],
            "05_non_screenable": results["non_screenable_records"],
            "06_manual_review": results["manual_review_needed"],
            "07_clean_master": results["clean_master_with_labels"],
            "08_database_stats": results["database_statistics"],
            "09_prisma_counts": results["prisma_counts"],
        },
        styled=True,
    )

    write_workbook_to_disk(
        title_path,
        {
            "Title_Screening_Results": results["title_screening_results"],
            "Title_Screening_Summary": results["title_screening_summary"],
        },
        styled=True,
    )

    write_workbook_to_disk(
        abstract_input_path,
        {
            "Abstract_Screening_Input": results["abstract_screening_input"],
        },
        styled=True,
    )

    return {
        "processing_path": processing_path,
        "title_path": title_path,
        "abstract_input_path": abstract_input_path,
    }