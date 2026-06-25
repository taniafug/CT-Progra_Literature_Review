import os
import re
import json
import time
from pathlib import Path
from typing import Dict, List, Any

import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI


load_dotenv()


# -----------------------------
# Default configuration
# -----------------------------
DEFAULT_MODEL = "gpt-4.1-mini"
DEFAULT_TITLE_COL = "Title"
DEFAULT_ABSTRACT_COL = "Abstract"

# -------------- This is just for knowing the charges-------------
MODEL_PRICES_PER_1M = {
    "gpt-4.1-mini": {
        "input": 0.40,
        "cached_input": 0.10,
        "output": 1.60,
    }
}

DEFAULT_SYSTEM_ROLE = "You are a strict screening assistant for a scoping review."

DEFAULT_PROMPT_INSTRUCTIONS = """
Classify the abstract strictly according to the screening rules and output schema.

GENERAL BEHAVIOR:
Use only information explicitly stated in the abstract. Do not invent information.
If the text is readable but lacks explicit evidence, code No.
Use Undetermined only when the abstract is incomplete, truncated, missing, or invalid.
Keep reasons concise.

EXTRA NOTES:
Not applicable is valid only for manualCheckReason when manualCheckNeeded = No.
For all screening criteria (q1, q2, q3), use only the allowed values in the Output Schema.
""".strip()

# -----------------------------
# Path utilities
# -----------------------------
def get_script_folder() -> Path:
    try:
        return Path(__file__).resolve().parent
    except NameError:
        return Path.cwd()


# -----------------------------
# Text normalization
# -----------------------------
def normalize_title(title: str) -> str:
    title = str(title).strip().lower()
    title = re.sub(r"\s+", " ", title)
    title = re.sub(r"[^\w\s]", "", title)
    return title.strip()


def is_missing_abstract(abstract: str) -> bool:
    value = str(abstract).strip().lower()
    return value in ["", "nan", "none", "null", "0", "0.0", "no abstract", "n/a", "na"]


# -----------------------------
# Duplicate handling
# -----------------------------
def remove_duplicates_keep_first(
    df: pd.DataFrame,
    title_col: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()
    df["Original_Row_Index"] = df.index
    df["Normalized_Title_For_Deduplication"] = df[title_col].apply(normalize_title)

    duplicate_mask = df.duplicated(
        subset=["Normalized_Title_For_Deduplication"],
        keep=False,
    )

    duplicate_rows = df[duplicate_mask].copy()
    summary_rows = []

    if not duplicate_rows.empty:
        for norm_title, group in duplicate_rows.groupby("Normalized_Title_For_Deduplication"):
            kept_original_index = group.iloc[0]["Original_Row_Index"]
            removed_original_indices = group.iloc[1:]["Original_Row_Index"].tolist()

            summary_rows.append(
                {
                    "Duplicated_Title": group.iloc[0][title_col],
                    "Normalized_Title": norm_title,
                    "Number_of_Occurrences": len(group),
                    "Kept_Original_Row_Index": kept_original_index,
                    "Removed_Original_Row_Indices": ", ".join(map(str, removed_original_indices)),
                }
            )

    duplicate_summary = pd.DataFrame(summary_rows)

    deduplicated_df = df.drop_duplicates(
        subset=["Normalized_Title_For_Deduplication"],
        keep="first",
    ).copy()

    deduplicated_df = deduplicated_df.drop(
        columns=["Normalized_Title_For_Deduplication"],
        errors="ignore",
    )

    return deduplicated_df, duplicate_summary


# -----------------------------
# Technical publication type helper
# -----------------------------
def detect_publication_type_from_text(title: str, abstract: str) -> tuple[str, str]:
    """
    Technical helper only. The LLM still classifies publicationTypeFlag using the rules file.
    This helper provides audit information and fallback grouping.
    """
    text = f"{title} {abstract}".lower()

    proceedings_patterns = [
        "proceedings",
        "conference proceedings",
        "front matter",
        "editorial",
        "preface",
        "table of contents",
        "book of abstracts",
        "conference program",
        "workshop proceedings",
        "symposium proceedings",
        "companion proceedings",
    ]

    poster_patterns = [
        "poster",
        "poster abstract",
        "poster presentation",
    ]

    conference_abstract_patterns = [
        "conference abstract",
        "extended abstract",
        "short paper",
        "workshop abstract",
        "work in progress",
        "wip",
        "doctoral consortium",
    ]

    matched_proceedings = [p for p in proceedings_patterns if p in text]
    matched_posters = [p for p in poster_patterns if p in text]
    matched_conference_abstracts = [p for p in conference_abstract_patterns if p in text]

    if matched_proceedings:
        return (
            "Proceedings-front matter",
            f"Technical keyword match: {', '.join(matched_proceedings)}",
        )

    if matched_posters:
        return (
            "Poster",
            f"Technical keyword match: {', '.join(matched_posters)}",
        )

    if matched_conference_abstracts:
        return (
            "Conference abstract",
            f"Technical keyword match: {', '.join(matched_conference_abstracts)}",
        )

    return "Undetermined", ""


# -----------------------------
# Read screening rules
# -----------------------------
def read_rules(excel_file) -> str:
    df = pd.read_excel(excel_file, sheet_name="Screening_Rules")
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Section", "Rule_Text"}
    missing = required - set(df.columns)

    if missing:
        raise ValueError(f"Missing required columns in Screening_Rules: {missing}")

    sections = []

    for section, group in df.groupby("Section", sort=False):
        rules = []

        for _, row in group.iterrows():
            rule = str(row["Rule_Text"]).strip()

            if rule and rule.lower() != "nan":
                rules.append(f"- {rule}")

        if rules:
            sections.append(f"{section}:\n" + "\n".join(rules))

    return "\n\n".join(sections)


# -----------------------------
# Read output schema
# -----------------------------
def read_output_schema(excel_file) -> pd.DataFrame:
    df = pd.read_excel(excel_file, sheet_name="Output_Schema")
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Field", "Description", "Allowed values", "Coding note"}
    missing = required - set(df.columns)

    if missing:
        raise ValueError(f"Missing required columns in Output_Schema: {missing}")

    df["Field"] = df["Field"].ffill()
    df["Description"] = df["Description"].ffill()
    df = df.dropna(subset=["Field"])

    return df


def build_schema_prompt(schema_df: pd.DataFrame) -> str:
    sections = []

    for field, group in schema_df.groupby("Field", sort=False):
        description = str(group["Description"].iloc[0]).strip()
        allowed_notes = []

        for _, row in group.iterrows():
            allowed = str(row["Allowed values"]).strip()
            note = str(row["Coding note"]).strip()

            if allowed.lower() in ["nan", ""]:
                continue

            text = f"- {allowed}"

            if note.lower() not in ["nan", ""]:
                text += f": {note}"

            allowed_notes.append(text)

        section = (
            f"FIELD: {field}\n"
            f"DESCRIPTION: {description}\n"
            f"ALLOWED VALUES / NOTES:\n"
            + "\n".join(allowed_notes)
        )

        sections.append(section)

    return "\n\n".join(sections)


def get_output_fields(schema_df: pd.DataFrame) -> List[str]:
    return list(
        dict.fromkeys(
            schema_df["Field"].dropna().astype(str).str.strip().tolist()
        )
    )


def make_json_schema(fields: List[str]) -> Dict[str, Any]:
    return {
        "type": "object",
        "properties": {field: {"type": "string"} for field in fields},
        "required": fields,
        "additionalProperties": False,
    }


# -----------------------------
# Read examples
# -----------------------------
def read_examples(excel_file) -> str:
    try:
        df = pd.read_excel(excel_file, sheet_name="Examples")
    except Exception:
        return ""

    if df.empty:
        return ""

    examples = []

    for _, row in df.iterrows():
        paper = str(row.get("Paper", "")).strip()
        abstract = str(row.get("Abstract", "")).strip()

        label = {}

        for col in df.columns:
            if col in ["Paper", "Abstract"]:
                continue

            value = row[col]

            if pd.isna(value):
                continue

            label[col] = str(value).strip()

        example_text = (
            f"Example paper:\n"
            f"Paper: {paper}\n"
            f"Abstract: {abstract}\n"
            f"Label:\n"
            f"{json.dumps(label, ensure_ascii=False, indent=2)}"
        )

        examples.append(example_text)

    return "\n\n".join(examples)

# ------------------------------------
# Read Adaptable Prompt
# -------------------------------------
def read_prompt_config(excel_file) -> dict:
    df = pd.read_excel(excel_file, sheet_name="Prompt_Config")
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Field", "Value"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns in Prompt_Config: {missing}")

    return {
        str(row["Field"]).strip(): str(row["Value"]).strip()
        for _, row in df.iterrows()
        if str(row["Field"]).strip()
    }

def build_default_prompt_instructions(prompt_config: dict) -> str:
    return f"""
{prompt_config.get("task_description", "")}

GENERAL BEHAVIOR:
{prompt_config.get("general_behavior", "")}

EXTRA NOTES:
{prompt_config.get("extra_notes", "")}
""".strip()

# -----------------------------
# Count inconsistent results in 3 runs
# -----------------------------

def build_decision_disagreement_summary(
    results_df: pd.DataFrame,
    number_of_runs: int,
) -> pd.DataFrame:
    rows = []

    for _, row in results_df.iterrows():
        decisions = []

        for run in range(1, number_of_runs + 1):
            decision = normalize_decision(row.get(f"Run_{run}_decision", ""))
            decisions.append(decision)

        unique_decisions = set(decisions)

        if len(unique_decisions) > 1:
            included_count = decisions.count("Included")
            excluded_count = decisions.count("Excluded")
            undetermined_count = decisions.count("Undetermined")
            other_count = decisions.count("Other")

            counts = {
                "Included": included_count,
                "Excluded": excluded_count,
                "Undetermined": undetermined_count,
                "Other": other_count,
            }

            majority_decision = max(counts, key=counts.get)

            minority_decisions = [
                decision for decision in unique_decisions
                if decision != majority_decision
            ]

            different_decision_count = sum(
                counts.get(decision, 0) for decision in minority_decisions
            )

            rows.append(
                {
                    "Paper": row.get("Paper", ""),
                    "Run_Decisions": ", ".join(decisions),
                    "Included_Count": included_count,
                    "Excluded_Count": excluded_count,
                    "Undetermined_Count": undetermined_count,
                    "Other_Count": other_count,
                    "Different_Decision_Count": different_decision_count,
                    "Different_Decision": ", ".join(minority_decisions),
                    "Final_Filter_Decision": row.get("Final_Filter_Decision", ""),
                }
            )

    return pd.DataFrame(rows)

# -----------------------------
# Abstract quality flag
# -----------------------------
def flag_abstract_quality(abstract: str) -> str:
    if is_missing_abstract(abstract):
        return "No valid abstract text available."

    text = str(abstract).strip()

    # Normalize invisible/control characters that often appear after PDF/Excel extraction
    text = text.replace("\r", "").replace("\n", " ").strip()

    # Common signs of corrupted or truncated extraction
    corruption_patterns = [
        r"\\$",                 # ends with backslash, e.g. Baden-W\
        r"\\\s*$",              # backslash followed by whitespace at the end
        r"\b[A-Z][a-z]+-W\\?$", # likely cut place/name, e.g. Baden-W\
        r"[�]{1,}",             # replacement character
        r"\\[a-zA-Z]+\{?$",     # broken LaTeX command
        r"\\'\{?[A-Za-z]?$",    # broken encoded character
    ]

    if any(re.search(pattern, text) for pattern in corruption_patterns):
        return "Abstract appears corrupted or truncated."

    # Detect likely incomplete ending
    last_words = text.split()[-5:]
    last_fragment = " ".join(last_words)

    incomplete_end_patterns = [
        r"\bfrom\s+[A-Z][A-Za-z-]*$",       # ends like "from six kindergartens in Baden"
        r"\bin\s+[A-Z][A-Za-z-]*\\?$",      # ends like "in Baden-W\"
        r"\bof\s+[A-Z][A-Za-z-]*$",
        r"\bthe\s+[A-Za-z-]+$",
        r"\band\s*$",
        r"\bor\s*$",
        r"\bwith\s*$",
        r"\bfrom\s*$",
        r"\bin\s*$",
    ]

    if any(re.search(pattern, last_fragment) for pattern in incomplete_end_patterns):
        return "Abstract appears truncated or ends with an incomplete phrase."

    if len(text.split()) < 30:
        return "Very short abstract; manual check recommended."

    if len(text.split()) < 60:
        return "Short abstract; interpret cautiously."

    return ""


def quality_note_requires_manual_check(quality_note: str) -> bool:
    note = str(quality_note).strip().lower()

    if not note:
        return False

    manual_indicators = [
        "no valid abstract",
        "corrupted",
        "truncated",
        "very short",
        "missing",
        "invalid",
    ]

    return any(indicator in note for indicator in manual_indicators)


def manual_review_requires_undetermined(quality_note: str, manual_check_needed: str, manual_check_reason: str) -> bool:
    """Return True when the record should not keep an Included/Excluded decision.

    In Stage 2, manual review means the abstract-level decision is not reliable enough
    to be treated as final. This is especially important for truncated, corrupted,
    missing, or very short abstracts: these should move forward as Undetermined,
    not as Excluded.
    """
    manual_value = str(manual_check_needed).strip().lower()
    if manual_value in ["yes", "yes for decision", "true", "1", "manual revision", "manual review"]:
        return True

    note = str(quality_note).strip().lower()
    reason = str(manual_check_reason).strip().lower()

    force_indicators = [
        "no valid abstract",
        "corrupted",
        "truncated",
        "very short",
        "missing",
        "invalid",
        "incomplete",
        "preventing reliable",
        "cannot reliably",
        "unable to reliably",
    ]

    combined = f"{note} {reason}"
    return any(indicator in combined for indicator in force_indicators)


def normalize_manual_check_fields(classification: Dict[str, Any]) -> Dict[str, Any]:
    manual_value = str(
        classification.get("manualCheckNeeded", "")
    ).strip().lower()

    reason = str(
        classification.get("manualCheckReason", "")
    ).strip()

    # Only normalize impossible states
    if manual_value == "no" and not reason:
        classification["manualCheckReason"] = "Not applicable"

    return classification

# -----------------------------
# Prompt
# -----------------------------
def build_prompt(
    paper: str,
    abstract: str,
    prompt_instructions: str,
    rules_prompt: str,
    schema_prompt: str,
    examples_prompt: str,
    technical_publication_type: str,
    technical_publication_reason: str,
) -> str:
    examples_section = ""

    if examples_prompt:
        examples_section = f"""
LABELLED EXAMPLES:
{examples_prompt}
"""

    return f"""
{prompt_instructions}

A technical keyword helper detected this possible publication type:
{technical_publication_type}
Reason:
{technical_publication_reason if technical_publication_reason else "No technical keyword match."}

Use this helper only as supporting information. Apply the screening rules and output schema as the main authority.

SCREENING RULES:
{rules_prompt}

OUTPUT SCHEMA:
{schema_prompt}

{examples_section}

PAPER TITLE:
{paper}

ABSTRACT:
{abstract}
""".strip()


# -----------------------------
# OpenAI call
# -----------------------------
def classify_abstract(
    client: OpenAI,
    model: str,
    paper: str,
    abstract: str,
    prompt_instructions: str,
    system_role: str,
    rules_prompt: str,
    schema_prompt: str,
    examples_prompt: str,
    json_schema: Dict[str, Any],
    technical_publication_type: str,
    technical_publication_reason: str,
) -> Dict[str, Any]:
    prompt = build_prompt(
        paper=paper,
        abstract=abstract,
        prompt_instructions=prompt_instructions,
        rules_prompt=rules_prompt,
        schema_prompt=schema_prompt,
        examples_prompt=examples_prompt,
        technical_publication_type=technical_publication_type,
        technical_publication_reason=technical_publication_reason,
    )

    response = client.responses.create(
        model=model,
        input=[
            {
                "role": "system",
                "content": system_role or DEFAULT_SYSTEM_ROLE,
            },
            {"role": "user", "content": prompt},
        ],
        temperature=0.0,
        text={
            "format": {
                "type": "json_schema",
                "name": "abstract_screening",
                "strict": True,
                "schema": json_schema,
            }
        },
    )

    classification = json.loads(response.output_text)

    usage = response.usage

    input_tokens = usage.input_tokens
    output_tokens = usage.output_tokens
    total_tokens = usage.total_tokens

    cached_input_tokens = 0
    if usage.input_tokens_details:
        cached_input_tokens = usage.input_tokens_details.cached_tokens or 0

    non_cached_input_tokens = input_tokens - cached_input_tokens

    prices = MODEL_PRICES_PER_1M.get(model, MODEL_PRICES_PER_1M["gpt-4.1-mini"])

    input_cost = (non_cached_input_tokens / 1_000_000) * prices["input"]
    cached_input_cost = (cached_input_tokens / 1_000_000) * prices["cached_input"]
    output_cost = (output_tokens / 1_000_000) * prices["output"]

    total_cost = input_cost + cached_input_cost + output_cost

    return {
        "classification": classification,
        "usage": {
            "input_tokens": input_tokens,
            "cached_input_tokens": cached_input_tokens,
            "output_tokens": output_tokens,
            "total_tokens": total_tokens,
            "input_cost_usd": input_cost,
            "cached_input_cost_usd": cached_input_cost,
            "output_cost_usd": output_cost,
            "total_cost_usd": total_cost,
        },
    }


# -----------------------------
# Final comparison logic
# -----------------------------
def normalize_decision(decision: str) -> str:
    decision = str(decision).strip().lower()

    if decision == "included":
        return "Included"

    if decision == "excluded":
        return "Excluded"

    if decision == "undetermined":
        return "Undetermined"

    return "Other"


def build_final_decision(row: pd.Series, number_of_runs: int) -> dict:
    decisions = []

    for run in range(1, number_of_runs + 1):
        decision_col = f"Run_{run}_decision"
        decisions.append(normalize_decision(row.get(decision_col, "")))

    included_count = decisions.count("Included")
    excluded_count = decisions.count("Excluded")
    undetermined_count = decisions.count("Undetermined")
    other_count = decisions.count("Other")

    # Conservative rule: any run that finds the record Undetermined keeps it for
    # manual/quality review rather than silently excluding it.
    if undetermined_count >= 1:
        final_decision = "Undetermined"

    elif included_count > excluded_count:
        final_decision = "Included"

    elif excluded_count > included_count:
        final_decision = "Excluded"

    # exact tie, no Undetermined present
    else:
        final_decision = "Undetermined"


    if included_count == number_of_runs:
        agreement_status = "Full inclusion agreement"
    elif excluded_count == number_of_runs:
        agreement_status = "Full exclusion agreement"
    elif undetermined_count == number_of_runs:
        agreement_status = "Full undetermined agreement"
    else:
        agreement_status = "Disagreement across runs"

    return {
        "Included_Count": included_count,
        "Excluded_Count": excluded_count,
        "Undetermined_Count": undetermined_count,
        "Other_Count": other_count,
        "Final_Filter_Decision": final_decision,
        "Agreement_Status": agreement_status,
    }


def get_run_values(row: pd.Series, number_of_runs: int, field: str) -> List[str]:
    values = []

    for run in range(1, number_of_runs + 1):
        col = f"Run_{run}_{field}"
        value = str(row.get(col, "")).strip()

        if value:
            values.append(value)

    return values


def most_common_value(values: List[str], default: str = "") -> str:
    cleaned = [str(v).strip() for v in values if str(v).strip()]

    if not cleaned:
        return default

    counts = {}
    for value in cleaned:
        counts[value] = counts.get(value, 0) + 1

    return max(counts, key=counts.get)


def get_final_field_value(
    row: pd.Series,
    number_of_runs: int,
    field: str,
    default: str = "",
) -> str:
    return most_common_value(get_run_values(row, number_of_runs, field), default=default)


def get_final_publication_type(row: pd.Series, number_of_runs: int) -> str:
    values = get_run_values(row, number_of_runs, "publicationTypeFlag")
    normalized = [str(v).strip() for v in values if str(v).strip()]

    if not normalized:
        return "Not specified"

    # Use majority when possible. If tied, prioritize clear non-full publication
    # types so they are not mixed into the conceptual screening pool.
    counts = {value: normalized.count(value) for value in set(normalized)}
    max_count = max(counts.values())
    tied = [value for value, count in counts.items() if count == max_count]

    priority = [
        "Proceedings-front matter",
        "Poster",
        "Conference abstract",
        "Undetermined",
        "Full paper",
        "Other",
        "Not specified",
    ]

    for item in priority:
        if item in tied:
            return item

    return tied[0]


def is_manual_revision_row(row: pd.Series, number_of_runs: int) -> bool:
    """
    Manual revision is reserved for records that cannot be reliably decided.
    Posters, conference abstracts, and proceedings/front matter can be excluded
    without manual review when their publication type is clear.
    """
    if str(row.get("Final_Filter_Decision", "")).strip() == "Undetermined":
        return True

    if quality_note_requires_manual_check(row.get("Abstract_Quality_Note", "")):
        return True

    for run in range(1, number_of_runs + 1):
        manual_col = f"Run_{run}_manualCheckNeeded"
        status_col = f"Run_{run}_Processing_Status"

        if str(row.get(manual_col, "")).strip().lower() == "yes":
            return True

        if str(row.get(status_col, "")).strip().lower() == "error":
            return True

    return False


def build_clean_summary_df(results_df: pd.DataFrame) -> pd.DataFrame:
    """Readable output without per-run token/cost/debug columns."""
    clean_columns = [
        "Paper",
        "Abstract",
        "Final_Filter_Decision",
        "Final_q1",
        "Final_q2",
        "Final_q3",
        "Final_Publication_Type_Flag",
        "Final_Reason",
        "Final_Confidence",
        "Final_ManualCheckNeeded",
        "Final_ManualCheckReason",
        "Agreement_Status",
        "Included_Count",
        "Excluded_Count",
        "Undetermined_Count",
        "Other_Count",
        "Abstract_Quality_Note",
    ]

    available_columns = [col for col in clean_columns if col in results_df.columns]
    return results_df[available_columns].copy() if available_columns else pd.DataFrame()


def build_usage_costs_df(results_df: pd.DataFrame, number_of_runs: int) -> pd.DataFrame:
    columns = [
        "Paper",
        "Abstract",
        "Final_Filter_Decision",
        "Final_Publication_Type_Flag",
        "Total_Runs_For_Paper",
        "Total_Input_Tokens",
        "Total_Output_Tokens",
        "Total_Tokens",
        "Total_Cost_USD",
    ]

    for run in range(1, number_of_runs + 1):
        columns.extend(
            [
                f"Run_{run}_Input_Tokens",
                f"Run_{run}_Cached_Input_Tokens",
                f"Run_{run}_Output_Tokens",
                f"Run_{run}_Total_Tokens",
                f"Run_{run}_Cost_USD",
                f"Run_{run}_Processing_Status",
                f"Run_{run}_Error_Message",
            ]
        )

    available_columns = [col for col in columns if col in results_df.columns]
    return results_df[available_columns].copy() if available_columns else pd.DataFrame()


def style_excel_workbook(path: Path) -> None:
    """Apply light styling to make workbook tabs easier to read."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    if not path.exists():
        return

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 80))
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 55))

        ws.auto_filter.ref = ws.dimensions

    wb.save(path)


# -----------------------------
# Excel output
# -----------------------------
def save_results_workbook(
    output_path: Path,
    results: List[Dict[str, Any]],
    duplicate_summary: pd.DataFrame | None = None,
    number_of_runs: int = 1,
    validation_sample_percent: int = 15,
) -> tuple[Path, Path]:
    results_df = pd.DataFrame(results)

    if duplicate_summary is None:
        duplicate_summary = pd.DataFrame()

    clean_df = build_clean_summary_df(results_df)

    if clean_df.empty:
        clean_df = pd.DataFrame()

    # Consistent final-decision groups
    included_df = clean_df[
        clean_df.get("Final_Filter_Decision", pd.Series(dtype=str))
        .astype(str).str.strip()
        == "Included"
    ].copy()

    excluded_df = clean_df[
        clean_df.get("Final_Filter_Decision", pd.Series(dtype=str))
        .astype(str).str.strip()
        == "Excluded"
    ].copy()

    undetermined_df = clean_df[
        clean_df.get("Final_Filter_Decision", pd.Series(dtype=str))
        .astype(str).str.strip()
        == "Undetermined"
    ].copy()

    manual_review_df = clean_df[
        clean_df.get("Final_ManualCheckNeeded", pd.Series(dtype=str))
        .astype(str).str.strip().str.lower()
        == "yes"
    ].copy()

    publication_type_excluded_df = clean_df[
        clean_df.get("Final_Publication_Type_Flag", pd.Series(dtype=str))
        .astype(str).str.strip()
        .isin(["Poster", "Conference abstract", "Proceedings-front matter"])
    ].copy()

    # Disagreements across runs
    decision_disagreement_df = build_decision_disagreement_summary(
        results_df,
        number_of_runs,
    )

    disagreement_papers = set(decision_disagreement_df.get("Paper", []))
    discrepancy_df = results_df[
        results_df["Paper"].isin(disagreement_papers)
    ].copy() if "Paper" in results_df.columns else pd.DataFrame()

    # Usage/costs
    usage_costs_df = build_usage_costs_df(results_df, number_of_runs)

    cost_summary_df = pd.DataFrame(
        [
            {"Metric": "Total records", "Value": len(results_df)},
            {"Metric": "Total API runs", "Value": len(results_df) * int(number_of_runs)},
            {"Metric": "Total input tokens", "Value": results_df.get("Total_Input_Tokens", pd.Series(dtype=float)).sum()},
            {"Metric": "Total output tokens", "Value": results_df.get("Total_Output_Tokens", pd.Series(dtype=float)).sum()},
            {"Metric": "Total tokens", "Value": results_df.get("Total_Tokens", pd.Series(dtype=float)).sum()},
            {"Metric": "Total cost USD", "Value": results_df.get("Total_Cost_USD", pd.Series(dtype=float)).sum()},
            {
                "Metric": "Average cost per record USD",
                "Value": (
                    results_df.get("Total_Cost_USD", pd.Series(dtype=float)).sum() / len(results_df)
                    if len(results_df) else 0
                ),
            },
        ]
    )

    decision_summary_df = pd.DataFrame(
        [
            {"Category": "Included", "Count": len(included_df)},
            {"Category": "Excluded", "Count": len(excluded_df)},
            {"Category": "Undetermined", "Count": len(undetermined_df)},
            {"Category": "Manual review needed", "Count": len(manual_review_df)},
            {"Category": "Publication-type excluded", "Count": len(publication_type_excluded_df)},
            {"Category": "Discrepancies across runs", "Count": len(discrepancy_df)},
            {"Category": "Duplicates removed", 
            "Count": (duplicate_summary["Number_of_Occurrences"]
            .sub(1).sum()
            if not duplicate_summary.empty
            else 0
            ),
            },
            {"Category": "Total", "Count": len(clean_df)},
        ]
    )

    # Validation sample
    validation_with_llm_df, manual_template_df, sampling_summary_df = build_validation_sample(
        results_df=results_df,
        number_of_runs=int(number_of_runs),
        sample_percent=int(validation_sample_percent),
    )

    # Workbook 1: clean readable results
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        decision_summary_df.to_excel(writer, sheet_name="Summary", index=False)
        included_df.to_excel(writer, sheet_name="Included", index=False)
        excluded_df.to_excel(writer, sheet_name="Excluded", index=False)
        undetermined_df.to_excel(writer, sheet_name="Undetermined", index=False)
        manual_review_df.to_excel(writer, sheet_name="Manual_Review", index=False)
        publication_type_excluded_df.to_excel(writer, sheet_name="Publication_Type_Excluded", index=False)
        clean_df.to_excel(writer, sheet_name="All_Clean_Summary", index=False)
        duplicate_summary.to_excel(writer, sheet_name="Duplicate_Summary", index=False)

    style_excel_workbook(output_path)

    # Workbook 2: audit, discrepancies, validation, costs
    audit_output_path = output_path.with_name(f"{output_path.stem}_audit_validation_usage.xlsx")

    with pd.ExcelWriter(audit_output_path, engine="openpyxl") as writer:
        # Put discrepancy first because it is the most important audit sheet
        decision_disagreement_df.to_excel(writer, sheet_name="Decision_Disagreements", index=False)
        discrepancy_df.to_excel(writer, sheet_name="Discrepant_Full_Runs", index=False)

        # Validation
        validation_with_llm_df.to_excel(writer, sheet_name="Validation_Sample_With_LLM", index=False)
        manual_template_df.to_excel(writer, sheet_name="Manual_Coding_Template", index=False)
        sampling_summary_df.to_excel(writer, sheet_name="Sampling_Summary", index=False)

        # Usage/costs
        cost_summary_df.to_excel(writer, sheet_name="Cost_Summary", index=False)
        usage_costs_df.to_excel(writer, sheet_name="Usage_Costs_Per_Run", index=False)

        # Full detailed audit
        results_df.to_excel(writer, sheet_name="Complete_Audit_All_Columns", index=False)
        duplicate_summary.to_excel(writer, sheet_name="Duplicate_Summary", index=False)

    style_excel_workbook(audit_output_path)

    return output_path, audit_output_path


# -----------------------------
# Validation process
# -----------------------------

def build_validation_sample(
    results_df: pd.DataFrame,
    number_of_runs: int,
    sample_percent: int = 15,
    random_state: int = 42,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:

    if results_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    df = results_df.copy()

    included_df = df[
        df["Final_Filter_Decision"].astype(str).str.strip() == "Included"
    ].copy()

    excluded_df = df[
        df["Final_Filter_Decision"].astype(str).str.strip() == "Excluded"
    ].copy()

    undetermined_df = df[
        df["Final_Filter_Decision"].astype(str).str.strip() == "Undetermined"
    ].copy()

    manual_df = df[
        df.apply(lambda row: is_manual_revision_row(row, number_of_runs), axis=1)
    ].copy()

    disagreement_df = df[
        df["Agreement_Status"].astype(str).str.strip() == "Disagreement across runs"
    ].copy()

    def sample_group(group_df: pd.DataFrame, group_name: str) -> pd.DataFrame:
        if group_df.empty:
            return pd.DataFrame()

        n = max(1, round(len(group_df) * sample_percent / 100))

        sampled = group_df.sample(
            n=min(n, len(group_df)),
            random_state=random_state,
        ).copy()

        sampled["Validation_Group"] = group_name
        return sampled

    sampled_included = sample_group(included_df, "Included")
    sampled_excluded = sample_group(excluded_df, "Excluded")
    sampled_undetermined = undetermined_df.copy()
    sampled_undetermined["Validation_Group"] = "Undetermined"

    sampled_manual = manual_df.copy()
    sampled_manual["Validation_Group"] = "Manual_Revision"

    sampled_disagreement = disagreement_df.copy()
    sampled_disagreement["Validation_Group"] = "Disagreement"

    validation_df = pd.concat(
        [
            sampled_included,
            sampled_excluded,
            sampled_undetermined,
            sampled_manual,
            sampled_disagreement,
        ],
        ignore_index=True,
    ).drop_duplicates(subset=["Paper", "Abstract"])

    llm_columns = [
        "Paper",
        "Abstract",
        "Validation_Group",
        "Abstract_Quality_Note",
        "Final_Filter_Decision",
        "Agreement_Status",
        "Final_Publication_Type_Flag",
        "Included_Count",
        "Excluded_Count",
        "Undetermined_Count",
        "Other_Count",
    ]

    for run in range(1, number_of_runs + 1):
        llm_columns.extend(
            [
                f"Run_{run}_q1",
                f"Run_{run}_q2",
                f"Run_{run}_q3",
                f"Run_{run}_decision",
                f"Run_{run}_confidence",
                f"Run_{run}_manualCheckNeeded",
                f"Run_{run}_manualCheckReason",
                f"Run_{run}_publicationTypeFlag",
                f"Run_{run}_reason",
            ]
        )

    llm_columns = [col for col in llm_columns if col in validation_df.columns]
    validation_with_llm_df = validation_df[llm_columns].copy()

    manual_template_df = validation_df[
        ["Paper", "Abstract", "Validation_Group"]
    ].copy()

    manual_template_df["Human_q1"] = ""
    manual_template_df["Human_q2"] = ""
    manual_template_df["Human_q3"] = ""
    manual_template_df["Human_decision"] = ""
    manual_template_df["Human_reason"] = ""
    manual_template_df["Human_confidence"] = ""

    sampling_summary_df = pd.DataFrame(
        [
            {
                "Group": "Included",
                "Total_records": len(included_df),
                "Sampled_records": len(sampled_included),
            },
            {
                "Group": "Excluded",
                "Total_records": len(excluded_df),
                "Sampled_records": len(sampled_excluded),
            },
            {
                "Group": "Undetermined",
                "Total_records": len(undetermined_df),
                "Sampled_records": len(sampled_undetermined),
            },
            {
                "Group": "Manual_Revision",
                "Total_records": len(manual_df),
                "Sampled_records": len(sampled_manual),
            },
            {
                "Group": "Disagreement",
                "Total_records": len(disagreement_df),
                "Sampled_records": len(sampled_disagreement),
            },
            {
                "Group": "Total unique validation sample",
                "Total_records": len(df),
                "Sampled_records": len(validation_with_llm_df),
            },
        ]
    )

    return validation_with_llm_df, manual_template_df, sampling_summary_df




# -----------------------------
# Organized workflow helpers
# -----------------------------
def load_screening_configuration(rules_file) -> Dict[str, Any]:
    """Read rules, schema, examples, prompt config, and JSON schema from the rules workbook."""
    rules_file.seek(0)
    rules_prompt = read_rules(rules_file)

    rules_file.seek(0)
    schema_df = read_output_schema(rules_file)
    schema_prompt = build_schema_prompt(schema_df)

    rules_file.seek(0)
    examples_prompt = read_examples(rules_file)

    rules_file.seek(0)
    prompt_config = read_prompt_config(rules_file)

    output_fields = get_output_fields(schema_df)
    json_schema = make_json_schema(output_fields)

    return {
        "rules_prompt": rules_prompt,
        "schema_df": schema_df,
        "schema_prompt": schema_prompt,
        "examples_prompt": examples_prompt,
        "prompt_config": prompt_config,
        "output_fields": output_fields,
        "json_schema": json_schema,
    }


def load_and_prepare_abstracts(
    abstract_file,
    title_col: str,
    abstract_col: str,
    deduplicate_titles: bool,
) -> tuple[pd.DataFrame, pd.DataFrame, int, int, int]:
    """Load the abstracts workbook, validate columns, and optionally remove duplicate titles."""
    df = pd.read_excel(abstract_file)
    df.columns = [str(c).strip() for c in df.columns]

    if title_col not in df.columns:
        raise ValueError(f"Title column '{title_col}' not found.")

    if abstract_col not in df.columns:
        raise ValueError(f"Abstract column '{abstract_col}' not found.")

    original_count = len(df)
    duplicate_summary = pd.DataFrame()

    if deduplicate_titles:
        df, duplicate_summary = remove_duplicates_keep_first(
            df=df,
            title_col=title_col,
        )

    deduplicated_count = len(df)
    removed_duplicates_count = original_count - deduplicated_count

    return df, duplicate_summary, original_count, deduplicated_count, removed_duplicates_count


def create_missing_abstract_classification(
    output_fields: List[str],
    quality_note: str,
) -> Dict[str, Any]:
    """Create a default Undetermined classification for missing or invalid abstracts."""
    classification = {field: "" for field in output_fields}
    classification["q1"] = "Undetermined"
    classification["q2"] = "Undetermined"
    classification["q3"] = "Undetermined"
    classification["decision"] = "Undetermined"
    classification["reason"] = "No valid abstract text available."
    classification["confidence"] = "Low"
    classification["manualCheckNeeded"] = "Yes"
    classification["manualCheckReason"] = quality_note or "No valid abstract text available."

    if "publicationTypeFlag" in output_fields:
        classification["publicationTypeFlag"] = "Undetermined"

    return classification


def classify_single_run(
    client: OpenAI,
    model: str,
    paper: str,
    abstract: str,
    prompt_instructions: str,
    system_role: str,
    rules_prompt: str,
    schema_prompt: str,
    examples_prompt: str,
    json_schema: Dict[str, Any],
    output_fields: List[str],
    technical_publication_type: str,
    technical_publication_reason: str,
    quality_note: str,
) -> tuple[Dict[str, Any], Dict[str, Any]]:
    """Classify one abstract once, or return a default result if the abstract is missing."""
    if is_missing_abstract(abstract):
        classification = create_missing_abstract_classification(output_fields, quality_note)
        usage = {
            "input_tokens": 0,
            "cached_input_tokens": 0,
            "output_tokens": 0,
            "total_tokens": 0,
            "total_cost_usd": 0,
        }
        return classification, usage

    llm_result = classify_abstract(
        client=client,
        model=model,
        paper=paper,
        abstract=abstract,
        prompt_instructions=prompt_instructions,
        system_role=system_role,
        rules_prompt=rules_prompt,
        schema_prompt=schema_prompt,
        examples_prompt=examples_prompt,
        json_schema=json_schema,
        technical_publication_type=technical_publication_type,
        technical_publication_reason=technical_publication_reason,
    )

    classification = llm_result["classification"]
    usage = llm_result["usage"]

    if quality_note_requires_manual_check(quality_note):
        classification["manualCheckNeeded"] = "Yes"
        classification["manualCheckReason"] = quality_note

    return classification, usage


def add_run_result_fields(
    result: Dict[str, Any],
    run: int,
    output_fields: List[str],
    classification: Dict[str, Any],
    usage: Dict[str, Any],
) -> Dict[str, Any]:
    """Add one successful run result to the paper result dictionary."""
    result[f"Run_{run}_Input_Tokens"] = usage["input_tokens"]
    result[f"Run_{run}_Cached_Input_Tokens"] = usage["cached_input_tokens"]
    result[f"Run_{run}_Output_Tokens"] = usage["output_tokens"]
    result[f"Run_{run}_Total_Tokens"] = usage["total_tokens"]
    result[f"Run_{run}_Cost_USD"] = usage["total_cost_usd"]

    classification = normalize_manual_check_fields(classification)

    for field in output_fields:
        result[f"Run_{run}_{field}"] = classification.get(field, "")

    result[f"Run_{run}_Processing_Status"] = "Success"
    result[f"Run_{run}_Error_Message"] = ""

    return result


def add_run_error_fields(
    result: Dict[str, Any],
    run: int,
    output_fields: List[str],
    error: Exception,
) -> Dict[str, Any]:
    """Add one failed run result to the paper result dictionary."""
    for field in output_fields:
        result[f"Run_{run}_{field}"] = ""

    result[f"Run_{run}_Input_Tokens"] = 0
    result[f"Run_{run}_Cached_Input_Tokens"] = 0
    result[f"Run_{run}_Output_Tokens"] = 0
    result[f"Run_{run}_Total_Tokens"] = 0
    result[f"Run_{run}_Cost_USD"] = 0
    result[f"Run_{run}_Processing_Status"] = "Error"
    result[f"Run_{run}_Error_Message"] = str(error)

    return result


def add_final_summary_fields(
    result: Dict[str, Any],
    number_of_runs: int,
    quality_note: str,
) -> Dict[str, Any]:
    """Add final decision, majority values, manual-check fields, and total usage/cost fields."""
    result_series = pd.Series(result)

    final_info = build_final_decision(
        row=result_series,
        number_of_runs=int(number_of_runs),
    )

    result.update(final_info)
    result_series = pd.Series(result)

    result["Final_q1"] = get_final_field_value(result_series, int(number_of_runs), "q1")
    result["Final_q2"] = get_final_field_value(result_series, int(number_of_runs), "q2")
    result["Final_q3"] = get_final_field_value(result_series, int(number_of_runs), "q3")
    result["Final_Reason"] = get_final_field_value(result_series, int(number_of_runs), "reason")
    result["Final_Confidence"] = get_final_field_value(result_series, int(number_of_runs), "confidence")
    result["Final_ManualCheckNeeded"] = get_final_field_value(
        result_series,
        int(number_of_runs),
        "manualCheckNeeded",
        default="No",
    )
    result["Final_ManualCheckReason"] = get_final_field_value(
        result_series,
        int(number_of_runs),
        "manualCheckReason",
        default="Not applicable",
    )

    result["Final_Publication_Type_Flag"] = get_final_publication_type(
        row=result_series,
        number_of_runs=int(number_of_runs),
    )

    # Keep the final manual-check fields aligned with the final decision and quality notes.
    if result["Final_Filter_Decision"] == "Undetermined" or quality_note_requires_manual_check(quality_note):
        result["Final_ManualCheckNeeded"] = "Yes"
        if not str(result.get("Final_ManualCheckReason", "")).strip() or str(result.get("Final_ManualCheckReason", "")).strip().lower() == "not applicable":
            result["Final_ManualCheckReason"] = quality_note or "Final decision is Undetermined."
    elif str(result.get("Final_ManualCheckNeeded", "")).strip().lower() == "no":
        result["Final_ManualCheckReason"] = "Not applicable"

    # Conservative scoping-review rule:
    # if a record needs manual review, the abstract-level decision should not remain
    # Included or Excluded. It should be treated as Undetermined until the reviewer
    # resolves it, so truncated abstracts do not become false exclusions.
    if manual_review_requires_undetermined(
        quality_note=quality_note,
        manual_check_needed=result.get("Final_ManualCheckNeeded", ""),
        manual_check_reason=result.get("Final_ManualCheckReason", ""),
    ):
        previous_decision = result.get("Final_Filter_Decision", "")
        if previous_decision != "Undetermined":
            result["Pre_Manual_Override_Decision"] = previous_decision
        result["Final_Filter_Decision"] = "Undetermined"
        result["Final_ManualCheckNeeded"] = "Yes"
        if not str(result.get("Final_ManualCheckReason", "")).strip() or str(result.get("Final_ManualCheckReason", "")).strip().lower() == "not applicable":
            result["Final_ManualCheckReason"] = quality_note or "Manual review required before final abstract decision."

    result["Total_Runs_For_Paper"] = int(number_of_runs)

    result["Total_Input_Tokens"] = sum(
        int(result.get(f"Run_{run}_Input_Tokens", 0) or 0)
        for run in range(1, int(number_of_runs) + 1)
    )

    result["Total_Output_Tokens"] = sum(
        int(result.get(f"Run_{run}_Output_Tokens", 0) or 0)
        for run in range(1, int(number_of_runs) + 1)
    )

    result["Total_Tokens"] = sum(
        int(result.get(f"Run_{run}_Total_Tokens", 0) or 0)
        for run in range(1, int(number_of_runs) + 1)
    )

    result["Total_Cost_USD"] = sum(
        float(result.get(f"Run_{run}_Cost_USD", 0) or 0)
        for run in range(1, int(number_of_runs) + 1)
    )

    return result


def classify_single_paper(
    client: OpenAI,
    model: str,
    paper: str,
    abstract: str,
    prompt_instructions: str,
    system_role: str,
    rules_prompt: str,
    schema_prompt: str,
    examples_prompt: str,
    json_schema: Dict[str, Any],
    output_fields: List[str],
    number_of_runs: int,
    log_callback=None,
    progress_callback=None,
    current_paper_index: int = 1,
    total_papers: int = 1,
    completed_iterations: int = 0,
    total_iterations: int = 1,
) -> tuple[Dict[str, Any], int]:
    """Classify one paper multiple times and return the final paper result."""
    quality_note = flag_abstract_quality(abstract)

    technical_publication_type, technical_publication_reason = detect_publication_type_from_text(
        title=paper,
        abstract=abstract,
    )

    result = {
        "Paper": paper,
        "Abstract": abstract,
        "Abstract_Quality_Note": quality_note,
    }

    for run in range(1, int(number_of_runs) + 1):
        if log_callback is not None:
            log_callback(
                f"Processing paper {current_paper_index}/{total_papers}, "
                f"run {run}/{number_of_runs}: "
                f"{paper[:80]}"
            )

        try:
            classification, usage = classify_single_run(
                client=client,
                model=model,
                paper=paper,
                abstract=abstract,
                prompt_instructions=prompt_instructions,
                system_role=system_role,
                rules_prompt=rules_prompt,
                schema_prompt=schema_prompt,
                examples_prompt=examples_prompt,
                json_schema=json_schema,
                output_fields=output_fields,
                technical_publication_type=technical_publication_type,
                technical_publication_reason=technical_publication_reason,
                quality_note=quality_note,
            )

            result = add_run_result_fields(
                result=result,
                run=run,
                output_fields=output_fields,
                classification=classification,
                usage=usage,
            )

        except Exception as e:
            result = add_run_error_fields(
                result=result,
                run=run,
                output_fields=output_fields,
                error=e,
            )

        completed_iterations += 1
        if progress_callback is not None:
            progress_callback(completed_iterations / total_iterations)

        time.sleep(0.3)

    result = add_final_summary_fields(
        result=result,
        number_of_runs=int(number_of_runs),
        quality_note=quality_note,
    )

    return result, completed_iterations



# ============================================================
# Checkpoint and duplicate-candidate helpers
# ============================================================

def make_processing_key(row: pd.Series, title_col: str = DEFAULT_TITLE_COL) -> str:
    """Build a stable key for checkpoint resume."""
    for col in ["Record_ID", "Paper_ID", "DOI", "DOI_Clean"]:
        value = str(row.get(col, "")).strip()
        if value and value.lower() not in ["nan", "none"]:
            return f"{col}:{value}"
    return "Title:" + normalize_title(row.get(title_col, ""))


def load_checkpoint_results(checkpoint_path: Path, title_col: str = DEFAULT_TITLE_COL) -> tuple[list[dict], set]:
    """Load checkpoint results and derive processed keys."""
    checkpoint_path = Path(checkpoint_path)
    if not checkpoint_path.exists():
        return [], set()

    try:
        checkpoint_df = pd.read_excel(checkpoint_path, dtype=str).fillna("")
    except Exception:
        return [], set()

    if checkpoint_df.empty:
        return [], set()

    results = checkpoint_df.to_dict("records")
    processed_keys = set()
    for _, row in checkpoint_df.iterrows():
        processed_keys.add(make_processing_key(row, title_col=title_col))
    return results, processed_keys


def text_similarity(a: Any, b: Any) -> float:
    """Compute normalized sequence similarity."""
    a = normalize_title(a)
    b = normalize_title(b)
    if not a or not b:
        return 0.0
    from difflib import SequenceMatcher
    return SequenceMatcher(None, a, b).ratio()


def normalize_doi_for_matching(value: Any) -> str:
    """Normalize DOI values for duplicate candidate detection."""
    if pd.isna(value):
        return ""
    text = str(value).lower().strip()
    text = text.replace("https://doi.org/", "")
    text = text.replace("http://doi.org/", "")
    text = text.replace("doi:", "")
    return text.strip().strip(".")


def first_author(value: Any) -> str:
    """Extract a normalized first-author approximation."""
    text = str(value or "").strip()
    if not text:
        return ""
    for sep in [" and ", ";", ","]:
        if sep in text:
            return normalize_title(text.split(sep)[0])
    return normalize_title(text)


def find_potential_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """Detect likely duplicate candidates after abstract input preparation."""
    if df is None or df.empty:
        return pd.DataFrame()

    work = df.copy()
    for col in ["Record_ID", "Title", "Abstract", "Authors", "Year", "DOI", "Source_Title", "Source_Database"]:
        if col not in work.columns:
            work[col] = ""

    work["_Title_Clean"] = work["Title"].apply(normalize_title)
    work["_Abstract_Clean"] = work["Abstract"].apply(normalize_title)
    work["_DOI_Clean"] = work["DOI"].apply(normalize_doi_for_matching)
    work["_First_Author"] = work["Authors"].apply(first_author)

    records = work.to_dict("records")
    rows = []

    for i in range(len(records)):
        for j in range(i + 1, len(records)):
            a = records[i]
            b = records[j]

            same_doi = bool(a["_DOI_Clean"]) and a["_DOI_Clean"] == b["_DOI_Clean"]
            same_year = str(a.get("Year", "")).strip() == str(b.get("Year", "")).strip()
            same_first_author = bool(a["_First_Author"]) and a["_First_Author"] == b["_First_Author"]

            title_score = text_similarity(a["_Title_Clean"], b["_Title_Clean"])
            abstract_score = text_similarity(a["_Abstract_Clean"], b["_Abstract_Clean"])

            likely_duplicate = (
                same_doi
                or title_score >= 0.96
                or abstract_score >= 0.94
                or (title_score >= 0.90 and abstract_score >= 0.85)
                or (same_year and same_first_author and title_score >= 0.86)
            )

            if not likely_duplicate:
                continue

            reason_parts = []
            if same_doi:
                reason_parts.append("same DOI")
            if same_year:
                reason_parts.append("same year")
            if same_first_author:
                reason_parts.append("same first author")
            if title_score >= 0.86:
                reason_parts.append("similar title")
            if abstract_score >= 0.85:
                reason_parts.append("similar abstract")

            rows.append({
                "Record_ID_A": a.get("Record_ID", ""),
                "Title_A": a.get("Title", ""),
                "Abstract_A": a.get("Abstract", ""),
                "Authors_A": a.get("Authors", ""),
                "Year_A": a.get("Year", ""),
                "DOI_A": a.get("DOI", ""),
                "Source_Title_A": a.get("Source_Title", ""),
                "Source_Database_A": a.get("Source_Database", ""),
                "Record_ID_B": b.get("Record_ID", ""),
                "Title_B": b.get("Title", ""),
                "Abstract_B": b.get("Abstract", ""),
                "Authors_B": b.get("Authors", ""),
                "Year_B": b.get("Year", ""),
                "DOI_B": b.get("DOI", ""),
                "Source_Title_B": b.get("Source_Title", ""),
                "Source_Database_B": b.get("Source_Database", ""),
                "Title_Similarity": round(title_score, 3),
                "Abstract_Similarity": round(abstract_score, 3),
                "Similarity_Reason": "; ".join(reason_parts),
                "LLM_Duplicate_Decision": "",
                "Recommended_Record_To_Keep": "",
                "Recommended_Record_To_Remove": "",
                "Conference_vs_Journal": "",
                "Manual_Review_Required": "Yes",
                "Rationale": "Potential duplicate detected from metadata and abstract similarity.",
            })

    return pd.DataFrame(rows)


def potential_duplicate_record_ids(potential_duplicates_df: pd.DataFrame | None) -> set:
    """Return all record IDs appearing in the duplicate candidate table."""
    if potential_duplicates_df is None or potential_duplicates_df.empty:
        return set()
    ids = set()
    if "Record_ID_A" in potential_duplicates_df.columns:
        ids.update(potential_duplicates_df["Record_ID_A"].astype(str).tolist())
    if "Record_ID_B" in potential_duplicates_df.columns:
        ids.update(potential_duplicates_df["Record_ID_B"].astype(str).tolist())
    return ids


def process_all_papers(
    df: pd.DataFrame,
    client: OpenAI,
    model: str,
    title_col: str,
    abstract_col: str,
    prompt_instructions: str,
    system_role: str,
    rules_prompt: str,
    schema_prompt: str,
    examples_prompt: str,
    json_schema: Dict[str, Any],
    output_fields: List[str],
    number_of_runs: int,
    checkpoint_path: Path,
    potential_duplicates_df: pd.DataFrame | None = None,
    log_callback=None,
    progress_callback=None,
) -> List[Dict[str, Any]]:
    """Process abstracts, save a checkpoint, and resume from existing checkpoint rows."""
    checkpoint_path = Path(checkpoint_path)
    results, processed_keys = load_checkpoint_results(checkpoint_path, title_col=title_col)
    duplicate_ids = potential_duplicate_record_ids(potential_duplicates_df)

    total_iterations = max(1, len(df) * int(number_of_runs))
    completed_iterations = min(len(results) * int(number_of_runs), total_iterations)

    for i, row_data in enumerate(df.to_dict("records"), start=1):
        row_series = pd.Series(row_data)
        row_key = make_processing_key(row_series, title_col=title_col)

        if row_key in processed_keys:
            if progress_callback is not None:
                progress_callback(min(completed_iterations / total_iterations, 1.0))
            continue

        paper = "" if pd.isna(row_series.get(title_col, "")) else str(row_series.get(title_col, "")).strip()
        abstract = "" if pd.isna(row_series.get(abstract_col, "")) else str(row_series.get(abstract_col, "")).strip()

        result, completed_iterations = classify_single_paper(
            client=client,
            model=model,
            paper=paper,
            abstract=abstract,
            prompt_instructions=prompt_instructions,
            system_role=system_role,
            rules_prompt=rules_prompt,
            schema_prompt=schema_prompt,
            examples_prompt=examples_prompt,
            json_schema=json_schema,
            output_fields=output_fields,
            number_of_runs=int(number_of_runs),
            log_callback=log_callback,
            progress_callback=progress_callback,
            current_paper_index=i,
            total_papers=len(df),
            completed_iterations=completed_iterations,
            total_iterations=total_iterations,
        )

        for col, value in row_data.items():
            if col not in result:
                result[col] = value

        record_id = str(row_data.get("Record_ID", "")).strip()
        result["Potential_Duplicate_Label"] = (
            "Potential duplicate – manual check" if record_id and record_id in duplicate_ids else ""
        )

        results.append(result)
        processed_keys.add(row_key)
        pd.DataFrame(results).to_excel(checkpoint_path, index=False)

    return results

def get_decision_group(clean_results_df: pd.DataFrame, decision: str) -> pd.DataFrame:
    """Return rows with a specific final decision."""
    if clean_results_df.empty:
        return pd.DataFrame()

    return clean_results_df[
        clean_results_df["Final_Filter_Decision"].astype(str).str.strip() == decision
    ].copy()


def get_manual_review_group(clean_results_df: pd.DataFrame) -> pd.DataFrame:
    """Return rows marked for manual review."""
    if clean_results_df.empty:
        return pd.DataFrame()

    return clean_results_df[
        clean_results_df.get("Final_ManualCheckNeeded", pd.Series(dtype=str))
        .astype(str)
        .str.strip()
        .str.lower()
        == "yes"
    ].copy()
