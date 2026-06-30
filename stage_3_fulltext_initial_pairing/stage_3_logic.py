import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import pypdf
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DECISION_VALUES = ["Included", "Excluded", "Undetermined"]


# ============================================================
# General utilities
# ============================================================
def reset_excel_pointer(excel_file) -> None:
    try:
        excel_file.seek(0)
    except Exception:
        pass


def safe_str(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def sanitize_text(text: str) -> str:
    text = re.sub(r"[\ud800-\udfff]", "", str(text))
    return text.encode("utf-8", "replace").decode("utf-8")


def truncate_text(text: str, max_chars: int) -> Tuple[str, str]:
    if len(text) <= max_chars:
        return text, "No"
    return text[:max_chars] + "\n\n[TEXT TRUNCATED FOR ANALYSIS]", "Yes"


def normalize_title(title: str) -> str:
    title = str(title).strip().lower()
    title = re.sub(r"\s+", " ", title)
    title = re.sub(r"[^\w\s]", "", title)
    return title.strip()


def normalize_decision(value: Any) -> str:
    value = safe_str(value)
    for allowed in DECISION_VALUES:
        if value.lower() == allowed.lower():
            return allowed
    return value


# ============================================================
# Rules template readers
# ============================================================
def read_rules(excel_file) -> str:
    reset_excel_pointer(excel_file)
    df = pd.read_excel(excel_file, sheet_name="Screening_Rules")
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Section", "Rule_Text"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns in Screening_Rules: {missing}")

    df["Section"] = df["Section"].ffill()
    df = df.dropna(subset=["Section", "Rule_Text"])

    sections = []
    for section, group in df.groupby("Section", sort=False):
        rules = []
        for _, row in group.iterrows():
            rule = safe_str(row["Rule_Text"])
            if rule:
                rules.append(f"- {rule}")
        if rules:
            sections.append(f"{section}:\n" + "\n".join(rules))

    return "\n\n".join(sections)


def read_output_schema(excel_file) -> pd.DataFrame:
    reset_excel_pointer(excel_file)
    df = pd.read_excel(excel_file, sheet_name="Output_Schema")
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Field", "Description", "Allowed values", "Coding note"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns in Output_Schema: {missing}")

    df["Field"] = df["Field"].ffill()
    df["Description"] = df["Description"].ffill()
    df = df.dropna(subset=["Field"])
    return df


def build_schema_prompt(schema_df: pd.DataFrame) -> str:
    sections = []

    for field, group in schema_df.groupby("Field", sort=False):
        description = safe_str(group["Description"].iloc[0])
        allowed_notes = []

        for _, row in group.iterrows():
            allowed = safe_str(row["Allowed values"])
            note = safe_str(row["Coding note"])
            if not allowed:
                continue

            text = f"- {allowed}"
            if note:
                text += f": {note}"
            allowed_notes.append(text)

        sections.append(
            f"FIELD: {field}\n"
            f"DESCRIPTION: {description}\n"
            f"ALLOWED VALUES / NOTES:\n"
            + "\n".join(allowed_notes)
        )

    return "\n\n".join(sections)


def get_output_fields(schema_df: pd.DataFrame) -> List[str]:
    return list(
        dict.fromkeys(
            schema_df["Field"].dropna().astype(str).str.strip().tolist()
        )
    )


def make_json_schema(fields: List[str]) -> Dict[str, Any]:
    return {
        "type": "object",
        "properties": {field: {"type": "string"} for field in fields},
        "required": fields,
        "additionalProperties": False,
    }


def read_prompt_config(excel_file) -> Dict[str, str]:
    reset_excel_pointer(excel_file)
    df = pd.read_excel(excel_file, sheet_name="Prompt_Config")
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Field", "Value"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns in Prompt_Config: {missing}")

    config: Dict[str, str] = {}
    for _, row in df.iterrows():
        key = safe_str(row.get("Field", ""))
        value = safe_str(row.get("Value", ""))
        if key and value:
            config[key] = value

    if not config:
        raise ValueError("Prompt_Config is empty. Add at least one Field/Value row.")

    return config


def build_prompt_instructions(prompt_config: Dict[str, str]) -> str:
    preferred_order = [
        "system_role",
        "task_description",
        "general_behavior",
        "input_text_scope",
        "extra_notes",
    ]

    sections: List[str] = []
    used = set()

    for key in preferred_order:
        if key in prompt_config and prompt_config[key].strip():
            sections.append(prompt_config[key].strip())
            used.add(key)

    for key, value in prompt_config.items():
        if key not in used and value.strip():
            sections.append(value.strip())

    return "\n\n".join(sections)


def read_examples(excel_file) -> str:
    reset_excel_pointer(excel_file)
    try:
        df = pd.read_excel(excel_file, sheet_name="Examples")
    except Exception:
        return ""

    if df.empty:
        return ""

    df.columns = [str(c).strip() for c in df.columns]
    examples = []

    for _, row in df.iterrows():
        paper = safe_str(row.get("Paper", ""))
        excerpt = safe_str(row.get("Full_Text_Excerpt", row.get("Abstract", "")))

        label = {}
        for col in df.columns:
            if col in ["Paper", "Full_Text_Excerpt", "Abstract"]:
                continue
            value = row[col]
            if pd.isna(value):
                continue
            label[col] = safe_str(value)

        examples.append(
            f"Example paper:\n"
            f"Paper: {paper}\n"
            f"Text excerpt: {excerpt}\n"
            f"Label:\n{json.dumps(label, ensure_ascii=False, indent=2)}"
        )

    return "\n\n".join(examples)


# ============================================================
# Section extraction
# ============================================================
def default_section_patterns() -> Dict[str, List[str]]:
    return {
        "introduction": ["introduction", "background"],
        "framework": [
            "theoretical framework",
            "conceptual framework",
            "computational thinking framework",
            "literature review",
            "related work",
            "review",
            "systematic review",
            "conceptual background",
        ],
        "methods": [
            "method",
            "methods",
            "methodology",
            "intervention",
            "procedure",
            "learning activities",
            "instructional design",
            "materials",
            "curriculum",
            "course design",
            "teaching activities",
            "implementation",
            "program design",
            "instruction",
            "learning design",
            "educational intervention",
            "assessment",
            "evaluation",
            "tasks",
            "activities",
            "research design",
            "data analysis",
        ],
        "results": ["results", "findings"],
        "discussion": ["discussion", "conclusion", "conclusions"],
        "appendix": ["appendix", "appendices", "supplementary material", "supplemental material"],
    }


def read_section_patterns(excel_file) -> Tuple[Dict[str, List[str]], str]:
    reset_excel_pointer(excel_file)
    try:
        df = pd.read_excel(excel_file, sheet_name="Section_Patterns")
    except Exception:
        return default_section_patterns(), "Technical fallback patterns used; Section_Patterns sheet not found."

    df.columns = [str(c).strip() for c in df.columns]
    required = {"Section_Group", "Heading"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns in Section_Patterns: {missing}")

    df = df.dropna(subset=["Section_Group", "Heading"])
    patterns: Dict[str, List[str]] = {}

    for _, row in df.iterrows():
        group = safe_str(row["Section_Group"]).lower()
        heading = safe_str(row["Heading"])
        if group and heading:
            patterns.setdefault(group, []).append(heading)

    if not patterns:
        raise ValueError("Section_Patterns sheet is empty.")

    return patterns, "Section_Patterns sheet used."


def find_section_positions(text: str, section_patterns: Dict[str, List[str]]) -> List[Tuple[str, int]]:
    positions = []

    for section_name, headings in section_patterns.items():
        for heading in headings:
            pattern = rf"(^|\n)\s*((\d+(\.\d+)*)|[IVX]+)?\.?\s*{re.escape(heading)}[:\s]*($|\n)"
            match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
            if match:
                positions.append((section_name, match.start()))
                break

    return sorted(positions, key=lambda x: x[1])


def extract_sections(text: str, section_patterns: Dict[str, List[str]]) -> Dict[str, str]:
    positions = find_section_positions(text, section_patterns)

    if not positions:
        return {"full_text_excerpt": text[:60000]}

    sections: Dict[str, str] = {}
    for idx, (section_name, start) in enumerate(positions):
        end = positions[idx + 1][1] if idx + 1 < len(positions) else len(text)
        sections[section_name] = text[start:end].strip()

    return sections


def build_relevant_text_for_screening(
    full_text: str,
    max_chars: int,
    section_patterns: Dict[str, List[str]],
) -> Tuple[str, str, str, str]:
    sections = extract_sections(full_text, section_patterns)
    found_sections = [s for s in sections if sections[s].strip()]
    section_detection_quality = "Good" if len(found_sections) >= 2 else "Weak"

    priority_sections = [
        "methods",
        "results",
        "framework",
        "discussion",
        "introduction",
        "appendix",
    ]

    selected_parts = []
    for section in priority_sections:
        if section in sections and sections[section].strip():
            selected_parts.append(f"\n\n===== {section.upper()} =====\n{sections[section]}")

    for section, content in sections.items():
        if section not in priority_sections and content.strip():
            selected_parts.append(f"\n\n===== {section.upper()} =====\n{content}")

    if not selected_parts:
        selected_parts.append(f"\n\n===== FULL TEXT EXCERPT =====\n{full_text[:max_chars]}")

    selected_text = "\n".join(selected_parts)
    selected_text, was_truncated = truncate_text(selected_text, max_chars)

    return selected_text, was_truncated, section_detection_quality, ", ".join(found_sections)


# ============================================================
# PDF reading
# ============================================================
def load_pdf_text(pdf_path: Path) -> str:
    try:
        reader = pypdf.PdfReader(str(pdf_path))
        pages = [page.extract_text() or "" for page in reader.pages]
        return sanitize_text("\n\n".join(pages))
    except Exception:
        return ""


# ============================================================
# Input dataframe helpers
# ============================================================
def remove_duplicates_keep_first(df: pd.DataFrame, title_col: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()
    df["Normalized_Title_For_Deduplication"] = df[title_col].apply(normalize_title)

    duplicate_mask = df.duplicated(
        subset=["Normalized_Title_For_Deduplication"],
        keep=False,
    )

    duplicate_rows = df[duplicate_mask].copy()
    summary_rows = []

    if not duplicate_rows.empty:
        for norm_title, group in duplicate_rows.groupby("Normalized_Title_For_Deduplication"):
            kept_index = group.index[0]
            removed_indices = group.index[1:].tolist()
            summary_rows.append(
                {
                    "Duplicated_Title": group.iloc[0][title_col],
                    "Normalized_Title": norm_title,
                    "Number_of_Occurrences": len(group),
                    "Kept_Row_Index": kept_index,
                    "Removed_Row_Indices": ", ".join(map(str, removed_indices)),
                }
            )

    duplicate_summary = pd.DataFrame(summary_rows)

    deduplicated_df = df.drop_duplicates(
        subset=["Normalized_Title_For_Deduplication"],
        keep="first",
    ).copy()

    deduplicated_df = deduplicated_df.drop(
        columns=["Normalized_Title_For_Deduplication"],
        errors="ignore",
    )

    return deduplicated_df, duplicate_summary


def prefix_input_columns(row_data: pd.Series) -> Dict[str, Any]:
    return {f"Input_{col}": safe_str(value) for col, value in row_data.items()}


# ============================================================
# Prompt + LLM classification
# ============================================================
def build_prompt(
    paper: str,
    selected_text: str,
    rules_prompt: str,
    schema_prompt: str,
    examples_prompt: str,
    text_was_truncated: str,
    section_detection_quality: str,
    prompt_instructions: str,
) -> str:
    examples_section = ""
    if examples_prompt:
        examples_section = f"""
LABELLED EXAMPLES:
{examples_prompt}
"""

    return f"""
{prompt_instructions}

TEXT METADATA:
- The selected text was truncated: {text_was_truncated}.
- Section detection quality: {section_detection_quality}.

If section detection quality is Weak, be cautious: relevant sections may be missing.
If this prevents reliable classification, use the appropriate Undetermined value and set manualCheckNeeded accordingly.

SCREENING RULES:
{rules_prompt}

OUTPUT SCHEMA:
{schema_prompt}

{examples_section}

PAPER TITLE:
{paper}

SELECTED FULL-TEXT SECTIONS:
{selected_text}
""".strip()


def extract_usage(response: Any) -> Dict[str, int]:
    usage = getattr(response, "usage", None)
    if usage is None:
        return {"input_tokens": 0, "output_tokens": 0, "total_tokens": 0}

    input_tokens = getattr(usage, "input_tokens", 0) or 0
    output_tokens = getattr(usage, "output_tokens", 0) or 0
    total_tokens = getattr(usage, "total_tokens", input_tokens + output_tokens) or (input_tokens + output_tokens)

    return {
        "input_tokens": int(input_tokens),
        "output_tokens": int(output_tokens),
        "total_tokens": int(total_tokens),
    }


def classify_fulltext(
    client,
    model: str,
    paper: str,
    selected_text: str,
    rules_prompt: str,
    schema_prompt: str,
    examples_prompt: str,
    json_schema: Dict[str, Any],
    text_was_truncated: str,
    section_detection_quality: str,
    prompt_instructions: str,
) -> Tuple[Dict[str, Any], Dict[str, int]]:
    prompt = build_prompt(
        paper=paper,
        selected_text=selected_text,
        rules_prompt=rules_prompt,
        schema_prompt=schema_prompt,
        examples_prompt=examples_prompt,
        text_was_truncated=text_was_truncated,
        section_detection_quality=section_detection_quality,
        prompt_instructions=prompt_instructions,
    )

    response = client.responses.create(
        model=model,
        input=[{"role": "user", "content": prompt}],
        text={
            "format": {
                "type": "json_schema",
                "name": "fulltext_screening",
                "strict": True,
                "schema": json_schema,
            }
        },
    )

    return json.loads(response.output_text), extract_usage(response)


# ============================================================
# Deterministic safeguards
# ============================================================
def append_manual_reason(current_reason: str, extra_reason: str) -> str:
    current_reason = safe_str(current_reason)
    extra_reason = safe_str(extra_reason)

    if not extra_reason:
        return current_reason if current_reason else "Not applicable"

    if current_reason.lower() in ["", "nan", "not applicable", "none", "null"]:
        return extra_reason

    return f"{current_reason} | {extra_reason}"


def enforce_manual_reason(classification: Dict[str, Any]) -> Dict[str, Any]:
    manual_needed = safe_str(classification.get("manualCheckNeeded", ""))
    if manual_needed == "No":
        classification["manualCheckReason"] = "Not applicable"
    return classification


def validate_decision_consistency(classification: Dict[str, Any]) -> Dict[str, Any]:
    qs = [
        safe_str(classification.get("q1", "")),
        safe_str(classification.get("q2", "")),
        safe_str(classification.get("q3", "")),
        safe_str(classification.get("q4", "")),
    ]

    q1, q2, q3, q4 = qs
    decision = safe_str(classification.get("decision", ""))

    if "Undetermined" in qs:
        expected = "Undetermined"
    elif q1 == "Yes" and q2 == "Yes" and q3 == "Yes" and q4 in ["Yes", "Partially"]:
        expected = "Included"
    else:
        expected = "Excluded"

    if decision != expected:
        classification["decision"] = expected
        classification["manualCheckNeeded"] = "Yes for decision"
        classification["manualCheckReason"] = append_manual_reason(
            classification.get("manualCheckReason", ""),
            f"Decision corrected from '{decision}' to '{expected}' based on q1–q4."
        )

    return classification


def postprocess_classification(classification: Dict[str, Any]) -> Dict[str, Any]:
    classification = enforce_manual_reason(classification)
    classification = validate_decision_consistency(classification)
    classification = enforce_manual_reason(classification)
    return classification


# ============================================================
# Output helpers
# ============================================================
def add_numbered_paper_id(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "Paper_ID" in df.columns:
        df = df.drop(columns=["Paper_ID"])
    df.insert(0, "Paper_ID", range(1, len(df) + 1))
    return df


def decision_summary(results_df: pd.DataFrame) -> pd.DataFrame:
    if results_df.empty or "decision" not in results_df.columns:
        return pd.DataFrame({"decision": DECISION_VALUES, "count": [0, 0, 0], "percentage": [0.0, 0.0, 0.0]})

    counts = (
        results_df["decision"]
        .map(normalize_decision)
        .value_counts()
        .reindex(DECISION_VALUES, fill_value=0)
        .reset_index()
    )
    counts.columns = ["decision", "count"]
    counts["percentage"] = (counts["count"] / max(1, len(results_df)) * 100).round(2)
    return counts


def build_prisma_summary(results_df: pd.DataFrame,duplicate_summary: Optional[pd.DataFrame] = None,no_full_text_count: int = 0,) -> pd.DataFrame:
    duplicate_count = 0
    if duplicate_summary is not None and not duplicate_summary.empty:
        if "Number_of_Occurrences" in duplicate_summary.columns:
            duplicate_count = int(
                pd.to_numeric(duplicate_summary["Number_of_Occurrences"], errors="coerce")
                .fillna(1)
                .sub(1)
                .clip(lower=0)
                .sum()
            )
        else:
            duplicate_count = len(duplicate_summary)

    if results_df.empty:
        return pd.DataFrame(
            [
                {"PRISMA_Metric": "Full-text records received from Stage 2", "Count": 0},
                {"PRISMA_Metric": "Safety duplicates removed before full-text screening", "Count": duplicate_count},
                {"PRISMA_Metric": "Full texts unavailable or not extractable", "Count": 0},
                {"PRISMA_Metric": "Full texts assessed by LLM", "Count": 0},
                {"PRISMA_Metric": "Full-text records included", "Count": 0},
                {"PRISMA_Metric": "Full-text records excluded", "Count": 0},
                {"PRISMA_Metric": "Full-text records undetermined", "Count": 0},
                {"PRISMA_Metric": "Full-text records requiring manual review", "Count": 0},
            ]
        )

    decision_clean = results_df.get("decision", pd.Series(dtype=str)).map(normalize_decision)
    processing = results_df.get("Processing_Status", pd.Series(dtype=str)).astype(str).str.strip()
    pdf_status = results_df.get("PDF_Status", pd.Series(dtype=str)).astype(str).str.strip()
    manual = results_df.get("manualCheckNeeded", pd.Series(dtype=str)).astype(str).str.strip().str.lower()

    unavailable_mask = (
        processing.eq("Skipped")
        | pdf_status.isin(["Missing PDF path", "PDF not found", "No extractable text"])
    )

    return pd.DataFrame(
        [
            {"PRISMA_Metric": "Full-text records received from Stage 2","Count": int(len(results_df)+ duplicate_count+ no_full_text_count)},
            {"PRISMA_Metric": "Safety duplicates removed before full-text screening", "Count": duplicate_count},
            {"PRISMA_Metric": "Full texts unavailable or not extractable", "Count": int(unavailable_mask.sum() + no_full_text_count)},
            {"PRISMA_Metric": "Full texts assessed by LLM", "Count": int(processing.eq("Success").sum())},
            {"PRISMA_Metric": "Full-text records included", "Count": int((decision_clean == "Included").sum())},
            {"PRISMA_Metric": "Full-text records excluded", "Count": int((decision_clean == "Excluded").sum())},
            {"PRISMA_Metric": "Full-text records undetermined", "Count": int((decision_clean == "Undetermined").sum())},
            {"PRISMA_Metric": "Full-text records requiring manual review", "Count": int(manual.isin(["yes", "yes for decision", "manual review", "manual revision", "true", "1"]).sum())},
        ]
    )


def random_sample_for_author_review(
    df: pd.DataFrame,
    fraction: float,
    random_seed: int,
    sample_group: str,
) -> pd.DataFrame:
    if df is None or df.empty or fraction <= 0:
        return pd.DataFrame()

    n = max(1, round(len(df) * fraction))
    n = min(n, len(df))

    sampled = df.sample(n=n, random_state=random_seed).copy()
    sampled.insert(0, "Sample_Group", sample_group)
    sampled.insert(1, "Sample_Size_Source_N", len(df))
    sampled.insert(2, "Sample_Percentage", fraction)

    return sampled


def author_sample_by_decision(results_df: pd.DataFrame, sample_fraction: float, random_seed: int) -> pd.DataFrame:
    if results_df.empty or "decision" not in results_df.columns:
        return pd.DataFrame()

    decision_clean = results_df["decision"].map(normalize_decision)

    included_df = results_df[decision_clean == "Included"].copy()
    excluded_df = results_df[decision_clean == "Excluded"].copy()
    undetermined_df = results_df[decision_clean == "Undetermined"].copy()

    sample_included = random_sample_for_author_review(
        included_df, sample_fraction, random_seed + 1, "Included random sample"
    )
    sample_excluded = random_sample_for_author_review(
        excluded_df, sample_fraction, random_seed + 2, "Excluded random sample"
    )
    sample_undetermined = random_sample_for_author_review(
        undetermined_df, sample_fraction, random_seed + 3, "Undetermined random sample"
    )

    samples = [df for df in [sample_included, sample_excluded, sample_undetermined] if not df.empty]
    if not samples:
        return pd.DataFrame()

    return pd.concat(samples, ignore_index=True)


def make_author_check_clean(author_sample: pd.DataFrame) -> pd.DataFrame:
    if author_sample.empty:
        return pd.DataFrame()

    cols = [c for c in ["Paper_ID", "Paper", "q1", "q2", "q3", "q4", "decision"] if c in author_sample.columns]
    clean = author_sample[cols].copy()

    clean["Author_q1"] = ""
    clean["Author_q2"] = ""
    clean["Author_q3"] = ""
    clean["Author_q4"] = ""
    clean["Author_decision"] = ""
    clean["Author_notes"] = ""

    return clean


def make_next_stage_included(results_df: pd.DataFrame) -> pd.DataFrame:
    """
    Creates the sheet for the next screening stage.

    This intentionally keeps only included papers, but preserves enough
    traceability to understand why each paper passed to the next stage.
    """
    cols = [
        "Paper_ID",
        "Paper",
        "summary",
        "PDF_path",
        "q1",
        "q2",
        "q3",
        "q4",
        "decision",
        "Matching_Evidence_Location",

    ]

    if results_df.empty or "decision" not in results_df.columns:
        return pd.DataFrame(columns=cols + ["Next_stage_label"])

    included = results_df[results_df["decision"].map(normalize_decision) == "Included"].copy()
    out = included[[c for c in cols if c in included.columns]].copy()

    if "PDF_path" not in out.columns and "Input_PDF_path" in included.columns:
        out["PDF_path"] = included["Input_PDF_path"]

    out["Next_stage_label"] = "Included for next stage"
    return out


def calculate_charges(
    results_df: pd.DataFrame,
    input_price_per_1m: float,
    output_price_per_1m: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    charges = results_df.copy()

    for col in ["input_tokens", "output_tokens", "total_tokens"]:
        if col not in charges.columns:
            charges[col] = 0
        charges[col] = pd.to_numeric(charges[col], errors="coerce").fillna(0)

    charges["input_cost"] = charges["input_tokens"] / 1_000_000 * input_price_per_1m
    charges["output_cost"] = charges["output_tokens"] / 1_000_000 * output_price_per_1m
    charges["total_cost"] = charges["input_cost"] + charges["output_cost"]

    keep_cols = [
        "Paper_ID",
        "Paper",
        "PDF_name",
        "Processing_Status",
        "input_tokens",
        "output_tokens",
        "total_tokens",
        "input_cost",
        "output_cost",
        "total_cost",
    ]
    charges = charges[[c for c in keep_cols if c in charges.columns]].copy()

    summary = pd.DataFrame(
        [
            {"metric": "input_tokens", "value": charges["input_tokens"].sum() if "input_tokens" in charges else 0},
            {"metric": "output_tokens", "value": charges["output_tokens"].sum() if "output_tokens" in charges else 0},
            {"metric": "total_tokens", "value": charges["total_tokens"].sum() if "total_tokens" in charges else 0},
            {"metric": "input_cost", "value": charges["input_cost"].sum() if "input_cost" in charges else 0},
            {"metric": "output_cost", "value": charges["output_cost"].sum() if "output_cost" in charges else 0},
            {"metric": "total_cost", "value": charges["total_cost"].sum() if "total_cost" in charges else 0},
        ]
    )

    return charges, summary


# ============================================================
# Excel style
# ============================================================
def autosize_worksheet(ws, max_width: int = 70) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = min(max(max(len(v) for v in values) + 2, 10), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        autosize_worksheet(ws)

    wb.save(path)

def get_no_full_text_count() -> int:
    output_dir = Path(__file__).resolve().parent / "output"
    candidate_paths = [
        output_dir / "stage_3_no_full_text_available.xlsx",
        output_dir / "no_full_text_available.xlsx",
    ]

    for no_full_text_path in candidate_paths:
        if not no_full_text_path.exists():
            continue

        try:
            no_full_text_df = pd.read_excel(
                no_full_text_path,
                dtype=str,
            )
            return len(no_full_text_df)

        except Exception:
            continue

    return 0

def write_checkpoint_excel(
    output_path: Path,
    results: List[Dict[str, Any]],
    duplicate_summary: Optional[pd.DataFrame] = None,
    sample_fraction: float = 0.15,
    random_seed: int = 42,
    input_price_per_1m: float = 0.40,
    output_price_per_1m: float = 1.60,
) -> Dict[str, pd.DataFrame]:
    results_df = pd.DataFrame(results)

    if not results_df.empty:
        results_df = add_numbered_paper_id(results_df)

    if "decision" in results_df.columns:
        results_df["decision"] = results_df["decision"].map(normalize_decision)

    summary_df = decision_summary(results_df)
    no_full_text_count = get_no_full_text_count()

    prisma_summary = build_prisma_summary(results_df,duplicate_summary,no_full_text_count=no_full_text_count,)

    author_sample = author_sample_by_decision(results_df, sample_fraction, random_seed)
    author_clean = make_author_check_clean(author_sample)
    next_stage = make_next_stage_included(results_df)
    charges_df, charges_summary = calculate_charges(results_df, input_price_per_1m, output_price_per_1m)

    decision_clean = results_df["decision"].map(normalize_decision) if "decision" in results_df.columns else pd.Series(dtype=str)

    included_df = results_df[decision_clean == "Included"].copy() if not results_df.empty else pd.DataFrame()
    excluded_df = results_df[decision_clean == "Excluded"].copy() if not results_df.empty else pd.DataFrame()
    undetermined_df = results_df[decision_clean == "Undetermined"].copy() if not results_df.empty else pd.DataFrame()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="All_results_corrected", index=False)
        summary_df.to_excel(writer, sheet_name="Decision_summary", index=False)
        prisma_summary.to_excel(writer, sheet_name="PRISMA_summary", index=False)
        included_df.to_excel(writer, sheet_name="Included", index=False)
        excluded_df.to_excel(writer, sheet_name="Excluded", index=False)
        undetermined_df.to_excel(writer, sheet_name="Undetermined", index=False)
        author_sample.to_excel(writer, sheet_name="Author_check_sample", index=False)
        author_clean.to_excel(writer, sheet_name="Author_check_clean", index=False)
        next_stage.to_excel(writer, sheet_name="Next_stage_included", index=False)
        charges_df.to_excel(writer, sheet_name="LLM_charges", index=False)
        charges_summary.to_excel(writer, sheet_name="LLM_charge_summary", index=False)

        if duplicate_summary is not None and not duplicate_summary.empty:
            duplicate_summary.to_excel(writer, sheet_name="Duplicates", index=False)

    style_workbook(output_path)

    return {
        "results": results_df,
        "summary": summary_df,
        "prisma_summary": prisma_summary,
        "included": included_df,
        "excluded": excluded_df,
        "undetermined": undetermined_df,
        "author_sample": author_sample,
        "author_clean": author_clean,
        "next_stage": next_stage,
        "charges": charges_df,
        "charges_summary": charges_summary,
    }


# ============================================================
# Final separated output workbooks
# ============================================================
def write_results_workbook(
    output_path: Path,
    results_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    prisma_summary: pd.DataFrame,
    included_df: pd.DataFrame,
    excluded_df: pd.DataFrame,
    undetermined_df: pd.DataFrame,
    next_stage_df: pd.DataFrame,
    duplicate_summary: Optional[pd.DataFrame] = None,
) -> None:
    """
    Main human-readable results workbook.

    Contains the full corrected results plus the sheet that goes directly
    to the next screening stage.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results_df.to_excel(writer, sheet_name="All_results_corrected", index=False)
        summary_df.to_excel(writer, sheet_name="Decision_summary", index=False)
        prisma_summary.to_excel(writer, sheet_name="PRISMA_summary", index=False)
        included_df.to_excel(writer, sheet_name="Included", index=False)
        excluded_df.to_excel(writer, sheet_name="Excluded", index=False)
        undetermined_df.to_excel(writer, sheet_name="Undetermined", index=False)
        next_stage_df.to_excel(writer, sheet_name="Next_stage_included", index=False)

        if duplicate_summary is not None and not duplicate_summary.empty:
            duplicate_summary.to_excel(writer, sheet_name="Duplicates", index=False)

    style_workbook(output_path)


def write_author_validation_workbook(
    output_path: Path,
    author_sample: pd.DataFrame,
    author_clean: pd.DataFrame,
) -> None:
    """
    Manual validation workbook.

    Author_check_sample keeps the sampled papers with context.
    Author_check_clean keeps only the compact fields to be completed manually.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        author_sample.to_excel(writer, sheet_name="Author_check_sample", index=False)
        author_clean.to_excel(writer, sheet_name="Author_check_clean", index=False)

    style_workbook(output_path)


def write_costs_workbook(
    output_path: Path,
    charges_df: pd.DataFrame,
    charges_summary: pd.DataFrame,
) -> None:
    """
    LLM token and cost workbook.
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        charges_df.to_excel(writer, sheet_name="LLM_charges", index=False)
        charges_summary.to_excel(writer, sheet_name="LLM_charge_summary", index=False)

    style_workbook(output_path)


def build_output_dataframes(results: List[Dict[str, Any]], duplicate_summary: Optional[pd.DataFrame] = None, sample_fraction: float = 0.15, random_seed: int = 42, input_price_per_1m: float = 0.40, output_price_per_1m: float = 1.60,) -> Dict[str, pd.DataFrame]:
    """
    Builds all output dataframes once, so they can be written into
    separated workbooks without recalculating inconsistently.
    """
    results_df = pd.DataFrame(results)

    if not results_df.empty:
        results_df = add_numbered_paper_id(results_df)

    if "decision" in results_df.columns:
        results_df["decision"] = results_df["decision"].map(normalize_decision)

    no_full_text_count = get_no_full_text_count()
    
    summary_df = decision_summary(results_df)
    prisma_summary = build_prisma_summary(results_df, duplicate_summary, no_full_text_count=no_full_text_count,)
    author_sample = author_sample_by_decision(results_df, sample_fraction, random_seed)
    author_clean = make_author_check_clean(author_sample)
    next_stage = make_next_stage_included(results_df)
    charges_df, charges_summary = calculate_charges(
        results_df,
        input_price_per_1m,
        output_price_per_1m,
    )

    decision_clean = (
        results_df["decision"].map(normalize_decision)
        if "decision" in results_df.columns
        else pd.Series(dtype=str)
    )

    included_df = results_df[decision_clean == "Included"].copy() if not results_df.empty else pd.DataFrame()
    excluded_df = results_df[decision_clean == "Excluded"].copy() if not results_df.empty else pd.DataFrame()
    undetermined_df = results_df[decision_clean == "Undetermined"].copy() if not results_df.empty else pd.DataFrame()

    return {
        "results": results_df,
        "summary": summary_df,
        "prisma_summary": prisma_summary,
        "included": included_df,
        "excluded": excluded_df,
        "undetermined": undetermined_df,
        "author_sample": author_sample,
        "author_clean": author_clean,
        "next_stage": next_stage,
        "charges": charges_df,
        "charges_summary": charges_summary,
    }


def write_output_workbooks(
    results_output_path: Path,
    validation_output_path: Path,
    costs_output_path: Path,
    results: List[Dict[str, Any]],
    duplicate_summary: Optional[pd.DataFrame] = None,
    sample_fraction: float = 0.15,
    random_seed: int = 42,
    input_price_per_1m: float = 0.40,
    output_price_per_1m: float = 1.60,
) -> Dict[str, pd.DataFrame]:
    """
    Writes the 3 final output files:

    1. Results workbook:
       - All_results_corrected
       - Decision_summary
       - Included
       - Excluded
       - Undetermined
       - Next_stage_included
       - Duplicates, when applicable

    2. Author/manual validation workbook:
       - Author_check_sample
       - Author_check_clean

    3. Costs workbook:
       - LLM_charges
       - LLM_charge_summary

    The technical checkpoint remains separate and can still be written
    with write_checkpoint_excel().
    """
    outputs = build_output_dataframes(
        results=results,
        duplicate_summary=duplicate_summary,
        sample_fraction=sample_fraction,
        random_seed=random_seed,
        input_price_per_1m=input_price_per_1m,
        output_price_per_1m=output_price_per_1m,
    )

    write_results_workbook(
        output_path=results_output_path,
        results_df=outputs["results"],
        summary_df=outputs["summary"],
        prisma_summary=outputs["prisma_summary"],
        included_df=outputs["included"],
        excluded_df=outputs["excluded"],
        undetermined_df=outputs["undetermined"],
        next_stage_df=outputs["next_stage"],
        duplicate_summary=duplicate_summary,
    )

    write_author_validation_workbook(
        output_path=validation_output_path,
        author_sample=outputs["author_sample"],
        author_clean=outputs["author_clean"],
    )

    write_costs_workbook(
        output_path=costs_output_path,
        charges_df=outputs["charges"],
        charges_summary=outputs["charges_summary"],
    )

    return outputs
